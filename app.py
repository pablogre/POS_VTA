# app.py - Sistema de Punto de Venta Argentina con Flask, MySQL, ARCA e Impresión Térmica

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import Numeric, or_, and_, func, desc, asc, case  
#from sqlalchemy import Numeric, or_, and_  # ← IMPORTAR AQUÍ
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
#import mysql.connector
from decimal import Decimal
from qr_afip import crear_generador_qr
import os
import requests
import xml.etree.ElementTree as ET
from zeep import Client
from zeep.wsse import BinarySignature
import base64
import hashlib
import csv
import io
from flask import make_response
from cryptography import x509
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import padding
import json
import subprocess
import MySQLdb.cursors
from estadisticas import init_estadisticas

# ================ FIX SSL COMPATIBLE PARA AFIP ================
import ssl
import urllib3
from urllib3.util import ssl_
from requests.adapters import HTTPAdapter
from requests import Session


def configurar_ssl_afip():
    """Configuración SSL compatible para todas las versiones de Python"""
    
    # Variables de entorno
    os.environ['PYTHONHTTPSVERIFY'] = '0'
    os.environ['CURL_CA_BUNDLE'] = ''
    os.environ['REQUESTS_CA_BUNDLE'] = ''
    
    def create_afip_ssl_context():
        """Crear contexto SSL para AFIP"""
        try:
            ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        except AttributeError:
            ctx = ssl.SSLContext(ssl.PROTOCOL_TLS)
        
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        # Ciphers más permisivos - probar niveles de seguridad
        try:
            ctx.set_ciphers('ALL:@SECLEVEL=0')
        except ssl.SSLError:
            try:
                ctx.set_ciphers('ALL:@SECLEVEL=1')
            except ssl.SSLError:
                ctx.set_ciphers('ALL')
        
        # Aplicar opciones SSL disponibles
        for opcion in ['OP_LEGACY_SERVER_CONNECT', 'OP_ALLOW_UNSAFE_LEGACY_RENEGOTIATION', 'OP_ALL']:
            if hasattr(ssl, opcion):
                try:
                    ctx.options |= getattr(ssl, opcion)
                except:
                    pass
        
        return ctx
    
    # Aplicar configuración
    ssl._create_default_https_context = create_afip_ssl_context
    ssl_.create_urllib3_context = create_afip_ssl_context
    urllib3.disable_warnings()
    
    print("✅ Configuración SSL para AFIP aplicada")

# Aplicar configuración SSL
configurar_ssl_afip()

def crear_session_afip():
    """Crear sesión HTTP personalizada para AFIP"""
    
    class AFIPAdapter(HTTPAdapter):
        def init_poolmanager(self, *args, **kwargs):
            try:
                ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
            except AttributeError:
                ctx = ssl.SSLContext(ssl.PROTOCOL_TLS)
            
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            try:
                ctx.set_ciphers('ALL:@SECLEVEL=0')
            except ssl.SSLError:
                try:
                    ctx.set_ciphers('ALL:@SECLEVEL=1')
                except ssl.SSLError:
                    ctx.set_ciphers('ALL')
            
            # Aplicar opciones disponibles
            for opcion in ['OP_LEGACY_SERVER_CONNECT', 'OP_ALLOW_UNSAFE_LEGACY_RENEGOTIATION', 'OP_ALL']:
                if hasattr(ssl, opcion):
                    try:
                        ctx.options |= getattr(ssl, opcion)
                    except:
                        pass
            
            kwargs['ssl_context'] = ctx
            return super().init_poolmanager(*args, **kwargs)
    
    session = Session()
    session.mount('https://', AFIPAdapter())
    session.verify = False
    return session

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui'

app.jinja_env.globals['hasattr'] = hasattr

# Intentar cargar configuración local, si no existe usar por defecto
try:
    from config_local import Config, ARCAConfig
    app.config.from_object(Config)
    ARCA_CONFIG = ARCAConfig()
except ImportError:
    # Configuración por defecto si no existe config_local.py
    app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://pos_user:pos_password@localhost/pos_argentina'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    
    class DefaultARCAConfig:
        CUIT = '20203852100'
        PUNTO_VENTA = 3
        CERT_PATH = 'certificados/certificado.crt'
        KEY_PATH = 'certificados/private.key'
        USE_HOMOLOGACION = False  # PRODUCCIÓN
        
        @property
        def WSAA_URL(self):
            return 'https://wsaahomo.afip.gov.ar/ws/services/LoginCms' if self.USE_HOMOLOGACION else 'https://wsaa.afip.gov.ar/ws/services/LoginCms'
        
        @property
        def WSFEv1_URL(self):
            return 'https://wswhomo.afip.gov.ar/wsfev1/service.asmx?WSDL' if self.USE_HOMOLOGACION else 'https://servicios1.afip.gov.ar/wsfev1/service.asmx?WSDL'
        
        TOKEN_CACHE_FILE = 'cache/token_arca.json'
    
    ARCA_CONFIG = DefaultARCAConfig()

db = SQLAlchemy(app)

# ================ SISTEMA DE IMPRESIÓN TÉRMICA ================
import tempfile
try:
    import win32print
    import win32api
    IMPRESION_DISPONIBLE = True
    print("✅ Sistema de impresión disponible")
except ImportError:
    IMPRESION_DISPONIBLE = False
    print("⚠️ Sistema de impresión no disponible (instalar: pip install pywin32)")

# IMPORTAR LA IMPRESORA DESDE EL ARCHIVO SEPARADO
from impresora_termica import impresora_termica

# Modelos de Base de Datos
class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    rol = db.Column(db.String(50), default='vendedor')
    activo = db.Column(db.Boolean, default=True)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    documento = db.Column(db.String(20))
    tipo_documento = db.Column(db.String(10))  # DNI, CUIT, etc.
    email = db.Column(db.String(100))
    telefono = db.Column(db.String(20))
    direccion = db.Column(db.Text)
    condicion_iva = db.Column(db.String(50))  # Responsable Inscripto, Monotributista, etc.

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    nombre = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    precio = db.Column(Numeric(10, 2), nullable=False)
    stock = db.Column(Numeric(10, 3), nullable=False, default=0.000)
    categoria = db.Column(db.String(100))
    iva = db.Column(Numeric(5, 2), default=21.00)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    costo = db.Column(Numeric(10, 2), default=0.00)  # ← AGREGAR ESTA LÍNEA
    margen = db.Column(Numeric(5, 2), default=30.00)  # ← AGREGAR ESTA LÍNEA
    
     # ✅ SOLO AGREGAR ESTOS 5 CAMPOS NUEVOS:
    es_combo = db.Column(db.Boolean, default=False)
    producto_base_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=True)
    cantidad_combo = db.Column(Numeric(8, 3), default=1.000)
    precio_unitario_base = db.Column(Numeric(10, 2), nullable=True)
    descuento_porcentaje = db.Column(Numeric(5, 2), default=0.00)
    acceso_rapido = db.Column(db.Boolean, default=False)  # ← AGREGAR ESTA LÍNEA
    orden_acceso_rapido = db.Column(db.Integer, default=0)  # ← Y ESTA LÍNEA TAMBIÉN

    # ✅ AGREGAR ESTA RELACIÓN:
    producto_base = db.relationship('Producto', remote_side=[id], backref='combos_derivados')
    
    def __repr__(self):
        return f'<Producto {self.codigo}: {self.nombre}>'
    
    # ✅ ACTUALIZAR tu método to_dict() existente:
    def to_dict(self):
        """Convertir producto a diccionario"""
        return {
            'id': self.id,
            'codigo': self.codigo,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'precio': float(self.precio),
            'costo': float(self.costo) if self.costo else 0.0,
            'margen': float(self.margen) if self.margen else 0.0,
            'stock': self.stock,
            'categoria': self.categoria,
            'iva': float(self.iva),
            'activo': self.activo,
            'fecha_creacion': self.fecha_creacion.isoformat() if self.fecha_creacion else None,
            'fecha_modificacion': self.fecha_modificacion.isoformat() if self.fecha_modificacion else None,
            
            # ✅ AGREGAR ESTAS LÍNEAS AL FINAL:
            'es_combo': self.es_combo,
            'producto_base_id': self.producto_base_id,
            'cantidad_combo': float(self.cantidad_combo) if self.cantidad_combo else 1.0,
            'precio_unitario_base': float(self.precio_unitario_base) if self.precio_unitario_base else float(self.precio),
            'descuento_porcentaje': float(self.descuento_porcentaje) if self.descuento_porcentaje else 0.0,
            'ahorro_combo': self.calcular_ahorro_combo(),
            'precio_normal': self.calcular_precio_normal(),
            'producto_base_nombre': self.producto_base.nombre if self.producto_base else None
        }
    
    @property
    def precio_calculado(self):
        """Calcular precio basado en costo y margen"""
        if self.costo and self.margen is not None:
            return float(self.costo) * (1 + (float(self.margen) / 100))
        return float(self.precio)
    
    def actualizar_precio_desde_costo_margen(self):
        """Actualizar el precio basado en costo y margen"""
        if self.costo and self.margen is not None:
            self.precio = Decimal(str(self.precio_calculado))
            self.fecha_modificacion = datetime.now()
    
    @staticmethod
    def calcular_precio_venta(costo, margen):
        """Método estático para calcular precio de venta"""
        if not costo or margen is None:
            return 0.0
        return float(costo) * (1 + (float(margen) / 100))
    # ✅ AGREGAR ESTOS MÉTODOS NUEVOS PARA COMBOS:
    def calcular_precio_normal(self):
        """Calcular precio normal sin descuento"""
        if self.es_combo and self.precio_unitario_base and self.cantidad_combo:
            return float(self.precio_unitario_base) * float(self.cantidad_combo)
        return float(self.precio)
    
    def calcular_ahorro_combo(self):
        """Calcular cuánto se ahorra con el combo"""
        if self.es_combo:
            precio_normal = self.calcular_precio_normal()
            precio_combo = float(self.precio)
            return precio_normal - precio_combo
        return 0.0
    
    def obtener_descripcion_completa(self):
        """Obtener descripción que incluye información del combo"""
        if self.es_combo:
            ahorro = self.calcular_ahorro_combo()
            cantidad_str = f"{self.cantidad_combo:g}"  # Elimina .0 si es entero
            return f"{self.nombre} - {cantidad_str} unidades (Ahorro: ${ahorro:.0f})"
        return self.nombre
    
    @staticmethod
    def obtener_productos_con_ofertas():
        """Obtener productos base con sus ofertas"""
        # Productos base (no combos)
        productos_base = Producto.query.filter_by(es_combo=False, activo=True).all()
        
        resultado = []
        for producto_base in productos_base:
            # Agregar producto base
            item_base = producto_base.to_dict()
            item_base['tipo'] = 'BASE'
            resultado.append(item_base)
            
            # Agregar sus combos/ofertas
            combos = Producto.query.filter_by(
                producto_base_id=producto_base.id, 
                es_combo=True, 
                activo=True
            ).order_by(Producto.precio).all()
            
            for combo in combos:
                item_combo = combo.to_dict()
                item_combo['tipo'] = 'COMBO'
                resultado.append(item_combo)
        
        return resultado

    # Agregar estos métodos a la clase Producto existente

    def obtener_precio_con_oferta(self, cantidad):
        """Obtener precio considerando ofertas por volumen"""
        try:
            cantidad_decimal = float(cantidad)
            
            # Buscar la mejor oferta aplicable
            oferta = OfertaVolumen.query.filter(
                and_(
                    OfertaVolumen.producto_id == self.id,
                    OfertaVolumen.cantidad_minima <= cantidad_decimal,
                    OfertaVolumen.activo == True
                )
            ).order_by(OfertaVolumen.cantidad_minima.desc()).first()
            
            if oferta:
                return float(oferta.precio_oferta)
            else:
                return float(self.precio)
                
        except Exception as e:
            print(f"Error calculando precio con oferta: {e}")
            return float(self.precio)

    def obtener_info_oferta(self, cantidad):
        """Obtener información detallada de la oferta aplicada"""
        try:
            cantidad_decimal = float(cantidad)
            precio_normal = float(self.precio)
            precio_con_oferta = self.obtener_precio_con_oferta(cantidad_decimal)
            
            # Buscar la oferta aplicada
            oferta = OfertaVolumen.query.filter(
                and_(
                    OfertaVolumen.producto_id == self.id,
                    OfertaVolumen.cantidad_minima <= cantidad_decimal,
                    OfertaVolumen.activo == True
                )
            ).order_by(OfertaVolumen.cantidad_minima.desc()).first()
            
            if oferta and precio_con_oferta < precio_normal:
                ahorro_unitario = precio_normal - precio_con_oferta
                ahorro_total = ahorro_unitario * cantidad_decimal
                
                return {
                    'tiene_oferta': True,
                    'precio_normal': precio_normal,
                    'precio_oferta': precio_con_oferta,
                    'ahorro_unitario': round(ahorro_unitario, 2),
                    'ahorro_total': round(ahorro_total, 2),
                    'cantidad_minima': float(oferta.cantidad_minima),
                    'descripcion_oferta': oferta.descripcion or f"Oferta por volumen desde {oferta.cantidad_minima} unidades"
                }
            
            return {
                'tiene_oferta': False,
                'precio_normal': precio_normal,
                'precio_oferta': precio_normal
            }
            
        except Exception as e:
            print(f"Error obteniendo info de oferta: {e}")
            return {
                'tiene_oferta': False,
                'precio_normal': float(self.precio),
                'precio_oferta': float(self.precio)
            }

    def tiene_ofertas_volumen(self):
        """Verificar si el producto tiene ofertas por volumen activas"""
        return OfertaVolumen.query.filter_by(
            producto_id=self.id,
            activo=True
        ).count() > 0

    @staticmethod
    def obtener_con_ofertas():
        """Obtener productos que tienen ofertas por volumen"""
        return db.session.query(Producto).join(OfertaVolumen).filter(
            and_(
                Producto.activo == True,
                OfertaVolumen.activo == True
            )
        ).distinct().all()    

class OfertaVolumen(db.Model):
    """Modelo para ofertas por volumen de productos"""
    __tablename__ = 'ofertas_volumen'
    
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cantidad_minima = db.Column(Numeric(10, 3), nullable=False)
    precio_oferta = db.Column(Numeric(10, 2), nullable=False)
    descripcion = db.Column(db.String(200))
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    # Relación con Producto
    producto = db.relationship('Producto', backref=db.backref('ofertas_volumen', lazy=True))
    
    def __repr__(self):
        return f'<OfertaVolumen {self.producto.codigo if self.producto else "SIN_PRODUCTO"}: {self.cantidad_minima}+ = ${self.precio_oferta}>'
    
    def to_dict(self):
        """Convertir a diccionario"""
        return {
            'id': self.id,
            'producto_id': self.producto_id,
            'cantidad_minima': float(self.cantidad_minima),
            'precio_oferta': float(self.precio_oferta),
            'descripcion': self.descripcion,
            'activo': self.activo,
            'fecha_creacion': self.fecha_creacion.isoformat() if self.fecha_creacion else None,
            'producto': {
                'codigo': self.producto.codigo,
                'nombre': self.producto.nombre
            } if self.producto else None
        }


    

class Factura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(50), unique=True)
    tipo_comprobante = db.Column(db.String(10))  # FA, FB, FC, etc.
    punto_venta = db.Column(db.Integer)
    fecha = db.Column(db.DateTime, default=datetime.now)  # ← Cambiar de utcnow a now
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    subtotal = db.Column(Numeric(10, 2))
    iva = db.Column(Numeric(10, 2))
    total = db.Column(Numeric(10, 2))
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, autorizada, anulada
    cae = db.Column(db.String(50))  # Código de Autorización Electrónico
    vto_cae = db.Column(db.Date)
    
    cliente = db.relationship('Cliente', backref='facturas')
    usuario = db.relationship('Usuario', backref='facturas')

class DetalleFactura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'))
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'))
    cantidad = db.Column(db.Integer, nullable=False)
    precio_unitario = db.Column(Numeric(10, 2), nullable=False)
    subtotal = db.Column(Numeric(10, 2), nullable=False)
    porcentaje_iva = db.Column(Numeric(5, 2), nullable=False, default=21.00)  # ← NUEVO CAMPO
    importe_iva = db.Column(Numeric(10, 2), nullable=False, default=0.00)    # ← NUEVO CAMPO
    
    factura = db.relationship('Factura', backref='detalles')
    producto = db.relationship('Producto', backref='detalles_factura')

class DescuentoFactura(db.Model):
    """Registro de descuentos aplicados a facturas - tabla independiente"""
    __tablename__ = 'descuentos_factura'
    
    id = db.Column(db.Integer, primary_key=True)
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'), unique=True, nullable=False)
    porcentaje_descuento = db.Column(Numeric(5, 2), nullable=False)
    monto_descuento = db.Column(Numeric(10, 2), nullable=False)
    total_original = db.Column(Numeric(10, 2), nullable=False)  # total antes del descuento
    fecha_aplicacion = db.Column(db.DateTime, default=datetime.now)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    # Relaciones
    factura = db.relationship('Factura', backref=db.backref('descuento_aplicado', uselist=False))
    usuario = db.relationship('Usuario', backref='descuentos_aplicados')
    
    def __repr__(self):
        return f'<DescuentoFactura {self.factura_id}: {self.porcentaje_descuento}% = ${self.monto_descuento}>'
    
    def to_dict(self):
        return {
            'id': self.id,
            'factura_id': self.factura_id,
            'porcentaje_descuento': float(self.porcentaje_descuento),
            'monto_descuento': float(self.monto_descuento),
            'total_original': float(self.total_original),
            'fecha_aplicacion': self.fecha_aplicacion.isoformat(),
            'usuario': self.usuario.nombre if self.usuario else None
        }

class MedioPago(db.Model):
    """Tabla para registrar los medios de pago de cada factura"""
    __tablename__ = 'medios_pago'
    
    id = db.Column(db.Integer, primary_key=True)
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'), nullable=False)
    medio_pago = db.Column(db.String(20), nullable=False)  # efectivo, credito, debito, mercado_pago
    importe = db.Column(Numeric(10, 2), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.now)
    
    # Relación con Factura
    factura = db.relationship('Factura', backref=db.backref('medios_pago', lazy=True))
    
    def __repr__(self):
        return f'<MedioPago {self.medio_pago}: ${self.importe}>'
    
    def to_dict(self):
        """Convertir a diccionario para JSON"""
        return {
            'id': self.id,
            'factura_id': self.factura_id,
            'medio_pago': self.medio_pago,
            'importe': float(self.importe),
            'fecha_registro': self.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
        }
    
    @staticmethod
    def obtener_medios_disponibles():
        """Retorna los medios de pago disponibles"""
        return [
            {'codigo': 'efectivo', 'nombre': 'Efectivo', 'icono': 'fas fa-money-bill-wave'},
            {'codigo': 'credito', 'nombre': 'Tarjeta de Crédito', 'icono': 'fas fa-credit-card'},
            {'codigo': 'debito', 'nombre': 'Tarjeta de Débito', 'icono': 'fas fa-credit-card'},
            {'codigo': 'mercado_pago', 'nombre': 'Mercado Pago', 'icono': 'fas fa-mobile-alt'}
        ]
    
    @staticmethod
    def calcular_recaudacion_por_fecha(fecha_desde, fecha_hasta):
        """Calcular recaudación por medio de pago en un rango de fechas"""
        try:
            from sqlalchemy import func, and_
            
            resultado = db.session.query(
                MedioPago.medio_pago,
                func.sum(MedioPago.importe).label('total'),
                func.count(MedioPago.id).label('cantidad_operaciones')
            ).filter(
                and_(
                    MedioPago.fecha_registro >= fecha_desde,
                    MedioPago.fecha_registro <= fecha_hasta
                )
            ).group_by(MedioPago.medio_pago).all()
            
            # Convertir a diccionario
            recaudacion = {}
            total_general = 0
            
            for medio, total, cantidad in resultado:
                recaudacion[medio] = {
                    'total': float(total),
                    'cantidad_operaciones': cantidad
                }
                total_general += float(total)
            
            return {
                'recaudacion_por_medio': recaudacion,
                'total_general': total_general,
                'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
                'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d')
            }
            
        except Exception as e:
            print(f"Error calculando recaudación: {e}")
            return None

# Agregar este modelo después de la clase MedioPago en tu app.py

class Gasto(db.Model):
    """Modelo para registrar gastos y egresos del negocio"""
    __tablename__ = 'gastos'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    descripcion = db.Column(db.Text, nullable=False)
    monto = db.Column(Numeric(10, 2), nullable=False)
    categoria = db.Column(db.String(50), nullable=False, default='general')
    metodo_pago = db.Column(db.String(30), nullable=False, default='efectivo')
    notas = db.Column(db.Text)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    activo = db.Column(db.Boolean, default=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))  # Usuario que registró el gasto
    
    # Relación con Usuario
    usuario = db.relationship('Usuario', backref=db.backref('gastos', lazy=True))
    
    def __repr__(self):
        return f'<Gasto {self.descripcion}: ${self.monto}>'
    
    def to_dict(self):
        """Convertir gasto a diccionario para JSON"""
        return {
            'id': self.id,
            'fecha': self.fecha.strftime('%Y-%m-%d') if self.fecha else None,
            'descripcion': self.descripcion,
            'monto': float(self.monto),
            'categoria': self.categoria,
            'metodo_pago': self.metodo_pago,
            'notas': self.notas,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None,
            'fecha_modificacion': self.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_modificacion else None,
            'activo': self.activo,
            'usuario': self.usuario.nombre if self.usuario else None
        }
    
    @staticmethod
    def obtener_categorias_disponibles():
        """Retorna las categorías de gastos disponibles"""
        return [
            {'codigo': 'general', 'nombre': 'General'},
            {'codigo': 'insumos', 'nombre': 'Insumos y Materiales'},
            {'codigo': 'servicios', 'nombre': 'Servicios (Luz, Gas, Internet)'},
            {'codigo': 'transporte', 'nombre': 'Transporte y Combustible'},
            {'codigo': 'personal', 'nombre': 'Gastos de Personal'},
            {'codigo': 'mantenimiento', 'nombre': 'Mantenimiento'},
            {'codigo': 'impuestos', 'nombre': 'Impuestos y Tasas'},
            {'codigo': 'otros', 'nombre': 'Otros'}
        ]
    
    @staticmethod
    def calcular_gastos_por_fecha(fecha_desde, fecha_hasta):
        """Calcular gastos por categoría en un rango de fechas"""
        try:
            resultado = db.session.query(
                Gasto.categoria,
                func.sum(Gasto.monto).label('total'),
                func.count(Gasto.id).label('cantidad_gastos')
            ).filter(
                and_(
                    Gasto.fecha >= fecha_desde,
                    Gasto.fecha <= fecha_hasta,
                    Gasto.activo == True
                )
            ).group_by(Gasto.categoria).all()
            
            # Convertir a diccionario
            gastos_por_categoria = {}
            total_general = 0
            
            for categoria, total, cantidad in resultado:
                gastos_por_categoria[categoria] = {
                    'total': float(total),
                    'cantidad_gastos': cantidad
                }
                total_general += float(total)
            
            return {
                'gastos_por_categoria': gastos_por_categoria,
                'total_general': total_general,
                'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
                'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d')
            }
            
        except Exception as e:
            print(f"Error calculando gastos: {e}")
            return None
    
    @staticmethod
    def obtener_gastos_por_medio_pago(fecha_desde, fecha_hasta):
        """Obtener gastos agrupados por medio de pago"""
        try:
            resultado = db.session.query(
                Gasto.metodo_pago,
                func.sum(Gasto.monto).label('total'),
                func.count(Gasto.id).label('cantidad')
            ).filter(
                and_(
                    Gasto.fecha >= fecha_desde,
                    Gasto.fecha <= fecha_hasta,
                    Gasto.activo == True
                )
            ).group_by(Gasto.metodo_pago).all()
            
            gastos_por_medio = {}
            for medio, total, cantidad in resultado:
                gastos_por_medio[medio] = {
                    'total': float(total),
                    'cantidad': cantidad
                }
            
            return gastos_por_medio
            
        except Exception as e:
            print(f"Error obteniendo gastos por medio: {e}")
            return {}
# ================== RUTAS API PARA REPORTES ==================


@app.route('/api/reporte_medios_pago')
def api_reporte_medios_pago():
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})

        desde = datetime.strptime(fecha_desde, "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        datos = MedioPago.calcular_recaudacion_por_fecha(desde, hasta)

        medios_pago = []
        if datos and 'recaudacion_por_medio' in datos:
            for medio, valores in datos['recaudacion_por_medio'].items():
                medios_pago.append({
                    'medio_pago': medio,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_operaciones']
                })

        return jsonify({
            'success': True,
            'reporte': {
                'medios_pago': medios_pago,
                'total_general': datos['total_general'] if datos else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/reporte_gastos')
def api_reporte_gastos():
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})

        desde = datetime.strptime(fecha_desde, "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        datos = Gasto.calcular_gastos_por_fecha(desde, hasta)

        gastos = []
        if datos and 'gastos_por_categoria' in datos:
            for categoria, valores in datos['gastos_por_categoria'].items():
                gastos.append({
                    'categoria': categoria,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_gastos']
                })

        return jsonify({
            'success': True,
            'reporte': {
                'gastos': gastos,
                'total_general': datos['total_general'] if datos else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/reporte_caja_diaria')
def api_reporte_caja_diaria():
    print(">>> API reporte_caja_diaria llamada")
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})

        desde = datetime.strptime(fecha_desde, "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        ingresos = MedioPago.calcular_recaudacion_por_fecha(desde, hasta)
        gastos = Gasto.calcular_gastos_por_fecha(desde, hasta)

        total_ingresos = ingresos['total_general'] if ingresos else 0
        total_gastos = gastos['total_general'] if gastos else 0
        balance = total_ingresos - total_gastos

        detalle_ingresos = []
        if ingresos and 'recaudacion_por_medio' in ingresos:
            for medio, valores in ingresos['recaudacion_por_medio'].items():
                detalle_ingresos.append({
                    'medio_pago': medio,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_operaciones']
                })

        detalle_gastos = []
        if gastos and 'gastos_por_categoria' in gastos:
            for categoria, valores in gastos['gastos_por_categoria'].items():
                detalle_gastos.append({
                    'categoria': categoria,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_gastos']
                })

        return jsonify({
            'success': True,
            'totalIngresos': total_ingresos,
            'totalGastos': total_gastos,
            'balance': balance,
            'detalleIngresos': detalle_ingresos,
            'detalleGastos': detalle_gastos
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# Clase para manejo de ARCA/AFIP
class ARCAClient:
    def __init__(self):
        self.config = ARCA_CONFIG
        self.token = None
        self.sign = None
        self.cuit = self.config.CUIT
        self.openssl_path = self._buscar_openssl()
        
        print(f"🔧 AFIP Client inicializado")
        print(f"   CUIT: {self.config.CUIT}")
        print(f"   Ambiente: {'HOMOLOGACIÓN' if self.config.USE_HOMOLOGACION else 'PRODUCCIÓN'}")
    
    def _buscar_openssl(self):
        """Buscar OpenSSL en ubicaciones conocidas"""
        ubicaciones = [
            './openssl.exe',
            'openssl.exe', 
            'openssl',
            r'C:\Program Files\OpenSSL-Win64\bin\openssl.exe',
            r'C:\OpenSSL-Win64\bin\openssl.exe',
            r'C:\Program Files (x86)\OpenSSL-Win32\bin\openssl.exe'
        ]
        
        for ubicacion in ubicaciones:
            try:
                if os.path.exists(ubicacion) or ubicacion in ['openssl.exe', 'openssl']:
                    result = subprocess.run([ubicacion, 'version'], 
                                          capture_output=True, text=True, timeout=5)
                    if result.returncode == 0:
                        print(f"✅ OpenSSL encontrado: {ubicacion}")
                        return ubicacion
            except:
                continue
        
        print("❌ OpenSSL no encontrado, usando 'openssl' por defecto")
        return 'openssl'
    
    def crear_tra(self):
        """Crear Ticket Request Access"""
        now = datetime.now()
        expire = now + timedelta(hours=12)
        unique_id = int(now.timestamp())
        
        tra_xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<loginTicketRequest version="1.0">
    <header>
        <uniqueId>{unique_id}</uniqueId>
        <generationTime>{now.strftime('%Y-%m-%dT%H:%M:%S.000-00:00')}</generationTime>
        <expirationTime>{expire.strftime('%Y-%m-%dT%H:%M:%S.000-00:00')}</expirationTime>
    </header>
    <service>wsfe</service>
</loginTicketRequest>'''
        
        return tra_xml
    
    def firmar_tra_openssl(self, tra_xml):
        """Firmar TRA usando OpenSSL"""
        try:
            import tempfile
            
            print(f"🔐 Firmando TRA con OpenSSL: {self.openssl_path}")
            
            # Verificar certificados
            if not os.path.exists(self.config.CERT_PATH):
                raise Exception(f"Certificado no encontrado: {self.config.CERT_PATH}")
            if not os.path.exists(self.config.KEY_PATH):
                raise Exception(f"Clave privada no encontrada: {self.config.KEY_PATH}")
            
            # Crear archivos temporales
            with tempfile.NamedTemporaryFile(mode='w', suffix='.xml', delete=False) as tra_file:
                tra_file.write(tra_xml)
                tra_temp = tra_file.name
            
            with tempfile.NamedTemporaryFile(suffix='.cms', delete=False) as cms_file:
                cms_temp = cms_file.name
            
            try:
                # Comando OpenSSL
                cmd = [
                    self.openssl_path, 'smime', '-sign',
                    '-in', tra_temp,
                    '-out', cms_temp,
                    '-signer', self.config.CERT_PATH,
                    '-inkey', self.config.KEY_PATH,
                    '-outform', 'DER',
                    '-nodetach'
                ]
                
                print(f"📝 Ejecutando firma...")
                
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
                
                if result.returncode != 0:
                    raise Exception(f"Error OpenSSL: {result.stderr}")
                
                # Leer archivo firmado
                with open(cms_temp, 'rb') as f:
                    cms_data = f.read()
                
                if len(cms_data) == 0:
                    raise Exception("Archivo CMS vacío")
                
                # Codificar en base64
                cms_b64 = base64.b64encode(cms_data).decode('utf-8')
                
                print("✅ TRA firmado correctamente")
                return cms_b64
                
            finally:
                # Limpiar archivos temporales
                try:
                    os.unlink(tra_temp)
                    os.unlink(cms_temp)
                except:
                    pass
                    
        except Exception as e:
            print(f"❌ Error firmando TRA: {e}")
            raise Exception(f"Error firmando TRA: {e}")
    
    def get_ticket_access(self):
        """Obtener ticket de acceso de WSAA con cache inteligente"""
        try:
            # *** NUEVO: Cache inteligente de tokens ***
            if hasattr(self, 'token_timestamp') and self.token and self.sign:
                # Verificar si el token aún es válido (duran 12 horas, usamos 10 horas para estar seguros)
                tiempo_transcurrido = datetime.now() - self.token_timestamp
                
                if tiempo_transcurrido < timedelta(hours=10):
                    print(f"🎫 Usando token existente (válido por {10 - tiempo_transcurrido.seconds//3600} horas más)")
                    return True
                else:
                    print("⏰ Token expirado, obteniendo uno nuevo...")
                    # Limpiar tokens viejos
                    self.token = None
                    self.sign = None
                    delattr(self, 'token_timestamp')
            
            print("🎫 Obteniendo nuevo ticket de acceso...")
            
            # Crear y firmar TRA
            tra_xml = self.crear_tra()
            tra_firmado = self.firmar_tra_openssl(tra_xml)
            
            # URL del WSAA
            wsaa_url = self.config.WSAA_URL + '?wsdl' if not self.config.WSAA_URL.endswith('?wsdl') else self.config.WSAA_URL
            
            print(f"🌐 Conectando con WSAA: {wsaa_url}")
            
            # USAR SESIÓN PERSONALIZADA
            session = crear_session_afip()
            
            from zeep.transports import Transport
            
            transport = Transport(session=session, timeout=60)
            client = Client(wsaa_url, transport=transport)
            
            # Enviar solicitud
            response = client.service.loginCms(tra_firmado)
            
            if response:
                # Procesar respuesta XML
                root = ET.fromstring(response)
                
                token_elem = root.find('.//token')
                sign_elem = root.find('.//sign')
                
                if token_elem is None or sign_elem is None:
                    raise Exception("Token o Sign no encontrados en respuesta")
                
                self.token = token_elem.text
                self.sign = sign_elem.text
                
                # *** NUEVO: Guardar timestamp del token ***
                self.token_timestamp = datetime.now()
                
                print("✅ Ticket de acceso obtenido y guardado en cache")
                return True
            else:
                raise Exception("Respuesta vacía de WSAA")
                
        except Exception as e:
            error_msg = str(e)
            
            # *** NUEVO: Manejo específico del error de token duplicado ***
            if "El CEE ya posee un TA valido" in error_msg:
                print("⚠️ AFIP indica que ya hay un token válido")
                print("💡 Esperando 30 segundos y reintentando...")
                
                import time
                time.sleep(30)  # Esperar 30 segundos
                
                # Limpiar tokens y reintentar UNA SOLA VEZ
                self.token = None
                self.sign = None
                if hasattr(self, 'token_timestamp'):
                    delattr(self, 'token_timestamp')
                
                # Reintentar una vez
                try:
                    print("🔄 Reintentando obtener token...")
                    tra_xml = self.crear_tra()
                    tra_firmado = self.firmar_tra_openssl(tra_xml)
                    
                    session = crear_session_afip()
                    transport = Transport(session=session, timeout=60)
                    client = Client(wsaa_url, transport=transport)
                    
                    response = client.service.loginCms(tra_firmado)
                    
                    if response:
                        root = ET.fromstring(response)
                        token_elem = root.find('.//token')
                        sign_elem = root.find('.//sign')
                        
                        if token_elem is not None and sign_elem is not None:
                            self.token = token_elem.text
                            self.sign = sign_elem.text
                            self.token_timestamp = datetime.now()
                            
                            print("✅ Token obtenido exitosamente en segundo intento")
                            return True
                    
                except Exception as e2:
                    print(f"❌ Segundo intento también falló: {e2}")
            
            print(f"❌ Error obteniendo ticket: {e}")
            return False

    def autorizar_comprobante(self, datos_comprobante):
        """
        Autorizar comprobante en AFIP usando WSFEv1 - VERSIÓN CORREGIDA CON MÚLTIPLES ALÍCUOTAS IVA
        """
        try:
            print("🎫 Verificando ticket de acceso...")
            
            # Verificar que tenemos ticket válido
            if not self.get_ticket_access():
                raise Exception("No se pudo obtener ticket de acceso")
            
            print("🌐 Conectando con WSFEv1...")
            
            # IMPORTANTE: URL limpia y configuración correcta
            wsfev1_url = 'https://servicios1.afip.gov.ar/wsfev1/service.asmx?WSDL'
            
            # Crear cliente SOAP con configuración específica
            session = crear_session_afip()
            from zeep.transports import Transport
            from zeep import Settings
            
            # Configuración específica para AFIP
            settings = Settings(strict=False, xml_huge_tree=True)
            transport = Transport(session=session, timeout=60, operation_timeout=60)
            
            try:
                from zeep import Client
                client = Client(wsfev1_url, transport=transport, settings=settings)
                print("✅ Cliente WSFEv1 creado correctamente")
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html', 'br line', 'span']):
                    raise Exception("WSFEv1 devolviendo HTML en lugar de XML - Servicio en mantenimiento")
                else:
                    raise Exception(f"Error creando cliente SOAP: {str(e)}")
            
            # Preparar autenticación
            auth = {
                'Token': self.token,
                'Sign': self.sign,
                'Cuit': self.cuit
            }
            
            print("📋 Preparando datos del comprobante...")
            
            # Obtener configuración
            pto_vta = datos_comprobante.get('punto_venta', self.config.PUNTO_VENTA)
            tipo_cbte = datos_comprobante.get('tipo_comprobante', 11)  # 11 = Factura C
            
            # Test rápido con FEDummy para verificar que el servicio funciona
            try:
                print("🧪 Verificando servicio con FEDummy...")
                dummy_response = client.service.FEDummy()
                print(f"✅ FEDummy OK: {dummy_response}")
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html']):
                    raise Exception("FEDummy devolviendo HTML - WSFEv1 en mantenimiento")
                else:
                    print(f"⚠️ Warning en FEDummy: {e}")
            
            # Obtener último comprobante autorizado
            try:
                print("📊 Consultando último comprobante autorizado...")
                ultimo_cbte_response = client.service.FECompUltimoAutorizado(
                    Auth=auth,
                    PtoVta=pto_vta,
                    CbteTipo=tipo_cbte
                )
                
                # Verificar errores en la respuesta
                if hasattr(ultimo_cbte_response, 'Errors') and ultimo_cbte_response.Errors:
                    print(f"⚠️ Advertencias al obtener último comprobante:")
                    if hasattr(ultimo_cbte_response.Errors, 'Err'):
                        errors = ultimo_cbte_response.Errors.Err
                        if isinstance(errors, list):
                            for error in errors:
                                print(f"   [{error.Code}] {error.Msg}")
                        else:
                            print(f"   [{errors.Code}] {errors.Msg}")
                
                ultimo_nro = getattr(ultimo_cbte_response, 'CbteNro', 0)
                proximo_nro = ultimo_nro + 1
                
                print(f"📊 Último comprobante AFIP: {ultimo_nro}")
                print(f"📊 Próximo número: {proximo_nro}")
                
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html']):
                    raise Exception("FECompUltimoAutorizado devolviendo HTML")
                else:
                    print(f"⚠️ Error obteniendo último comprobante: {e}")
                    print("🔄 Usando número secuencial local...")
                    proximo_nro = 1
            
            # Preparar datos del comprobante
            fecha_hoy = datetime.now().strftime('%Y%m%d')
            
            # *** NUEVO: CALCULAR ALÍCUOTAS IVA SEPARADAS ***
            # Obtener los items del comprobante (deben venir con detalle por producto)
            items_detalle = datos_comprobante.get('items_detalle', [])
            
            if not items_detalle:
                raise Exception("Se requieren items detallados con alícuotas IVA individuales")
            
            # Agrupar por alícuota de IVA
            alicuotas_iva = {}
            importe_neto_total = 0
            importe_iva_total = 0
            
            print("🧮 Calculando alícuotas de IVA por separado...")
            
            for item in items_detalle:
                subtotal = float(item.get('subtotal', 0))
                iva_porcentaje = float(item.get('iva_porcentaje', 0))
                
                # Calcular IVA del item con redondeo AFIP
                iva_item = round((subtotal * iva_porcentaje / 100), 2)
                
                # Agrupar por alícuota
                if iva_porcentaje not in alicuotas_iva:
                    alicuotas_iva[iva_porcentaje] = {
                        'base_imponible': 0,
                        'iva_total': 0
                    }
                
                alicuotas_iva[iva_porcentaje]['base_imponible'] += subtotal
                alicuotas_iva[iva_porcentaje]['iva_total'] += iva_item
                
                importe_neto_total += subtotal
                importe_iva_total += iva_item
                
                print(f"   📦 Item: ${subtotal:.2f} (IVA {iva_porcentaje}% = ${iva_item:.2f})")
            
            # Redondear totales
            importe_neto_total = round(importe_neto_total, 2)
            importe_iva_total = round(importe_iva_total, 2)
            importe_total = round(importe_neto_total + importe_iva_total, 2)
            
            print(f"💰 Totales calculados: Neto=${importe_neto_total:.2f}, IVA=${importe_iva_total:.2f}, Total=${importe_total:.2f}")
            
            # Mostrar alícuotas calculadas
            print("📊 Alícuotas de IVA:")
            for porcentaje, datos in alicuotas_iva.items():
                base = round(datos['base_imponible'], 2)
                iva = round(datos['iva_total'], 2)
                print(f"   IVA {porcentaje}%: Base=${base:.2f}, IVA=${iva:.2f}")
            
            # Estructura del comprobante según especificación AFIP
            comprobante = {
                'Concepto': 1,
                'DocTipo': datos_comprobante.get('doc_tipo', 99),
                'DocNro': datos_comprobante.get('doc_nro', 0),
                'CbteDesde': proximo_nro,
                'CbteHasta': proximo_nro,
                'CbteFch': fecha_hoy,
                'ImpTotal': importe_total,
                'ImpTotConc': 0.00,
                'ImpNeto': importe_neto_total,
                'ImpOpEx': 0.00,
                'ImpTrib': 0.00,
                'ImpIVA': importe_iva_total,
                'MonId': 'PES',
                'MonCotiz': 1.00,
            }
            
            # *** CLAVE: AGREGAR DETALLE DE IVA POR ALÍCUOTA ***
            if importe_iva_total > 0:
                alicuotas_afip = []
                
                for porcentaje, datos in alicuotas_iva.items():
                    if porcentaje > 0:  # Solo agregar si hay IVA
                        # Mapear porcentajes a códigos AFIP
                        codigo_iva = self.get_codigo_iva_afip(porcentaje)
                        
                        if codigo_iva:
                            alicuotas_afip.append({
                                'Id': codigo_iva,
                                'BaseImp': round(datos['base_imponible'], 2),
                                'Importe': round(datos['iva_total'], 2)
                            })
                            
                            print(f"✅ Alícuota AFIP: Código {codigo_iva}, Base=${datos['base_imponible']:.2f}, IVA=${datos['iva_total']:.2f}")
                
                if alicuotas_afip:
                    comprobante['Iva'] = {'AlicIva': alicuotas_afip}
                    print(f"📝 Se agregaron {len(alicuotas_afip)} alícuotas de IVA al comprobante")
                else:
                    print("⚠️ No se pudieron mapear las alícuotas a códigos AFIP")
            
            # Crear request completo
            fe_request = {
                'FeCabReq': {
                    'CantReg': 1,
                    'PtoVta': pto_vta,
                    'CbteTipo': tipo_cbte
                },
                'FeDetReq': {
                    'FECAEDetRequest': [comprobante]
                }
            }
            
            print("📤 Enviando solicitud de autorización a AFIP...")
            print(f"   Tipo comprobante: {tipo_cbte}")
            print(f"   Punto de venta: {pto_vta}")
            print(f"   Número: {proximo_nro}")
            print(f"   Fecha: {fecha_hoy}")
            print(f"   Total: ${importe_total:.2f}")
            print(f"   Alícuotas IVA: {len(alicuotas_iva)} diferentes")
            
            # ENVÍO CRÍTICO
            try:
                response = client.service.FECAESolicitar(Auth=auth, FeCAEReq=fe_request)
                print("✅ Respuesta recibida de AFIP")
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html', 'br line', 'span']):
                    raise Exception("FECAESolicitar devolviendo HTML - WSFEv1 en mantenimiento")
                else:
                    raise Exception(f"Error en FECAESolicitar: {str(e)}")
            
            # Procesar respuesta de AFIP
            print("📋 Procesando respuesta de AFIP...")
            
            # Verificar errores generales
            if hasattr(response, 'Errors') and response.Errors:
                errores = []
                if hasattr(response.Errors, 'Err'):
                    errors = response.Errors.Err
                    if isinstance(errors, list):
                        for error in errors:
                            errores.append(f"[{error.Code}] {error.Msg}")
                    else:
                        errores.append(f"[{errors.Code}] {errors.Msg}")
                
                error_msg = " | ".join(errores)
                raise Exception(f"Errores AFIP: {error_msg}")
            
            # Verificar que hay respuesta de detalle
            if not hasattr(response, 'FeDetResp') or not response.FeDetResp:
                raise Exception("Respuesta de AFIP sin detalles")
            
            # Obtener detalle de respuesta
            if not hasattr(response.FeDetResp, 'FECAEDetResponse'):
                raise Exception("Respuesta de AFIP sin FECAEDetResponse")
            
            detalle_resp = response.FeDetResp.FECAEDetResponse[0]
            
            # Verificar resultado
            resultado = getattr(detalle_resp, 'Resultado', None)
            if resultado != 'A':  # A = Aprobado
                observaciones = []
                if hasattr(detalle_resp, 'Observaciones') and detalle_resp.Observaciones:
                    if hasattr(detalle_resp.Observaciones, 'Obs'):
                        obs_list = detalle_resp.Observaciones.Obs
                        if isinstance(obs_list, list):
                            for obs in obs_list:
                                observaciones.append(f"[{obs.Code}] {obs.Msg}")
                        else:
                            observaciones.append(f"[{obs_list.Code}] {obs_list.Msg}")
                
                obs_msg = " | ".join(observaciones) if observaciones else "Sin observaciones"
                raise Exception(f"Comprobante no autorizado. Resultado: {resultado}. {obs_msg}")
            
            # ÉXITO - Extraer datos
            cae = getattr(detalle_resp, 'CAE', None)
            fecha_vencimiento = getattr(detalle_resp, 'CAEFchVto', None)
            
            if not cae:
                raise Exception("Respuesta sin CAE")
            
            if not fecha_vencimiento:
                raise Exception("Respuesta sin fecha de vencimiento CAE")
            
            numero_completo = f"{pto_vta:04d}-{proximo_nro:08d}"
            
            print(f"🎉 ¡COMPROBANTE AUTORIZADO EXITOSAMENTE!")
            print(f"   Número: {numero_completo}")
            print(f"   CAE: {cae}")
            print(f"   Vencimiento CAE: {fecha_vencimiento}")
            
            return {
                'success': True,
                'cae': cae,
                'numero': numero_completo,
                'punto_venta': pto_vta,
                'numero_comprobante': proximo_nro,
                'fecha_vencimiento': fecha_vencimiento,
                'fecha_proceso': fecha_hoy,
                'importe_total': importe_total,
                'tipo_comprobante': tipo_cbte,
                'estado': 'autorizada',
                'vto_cae': datetime.strptime(fecha_vencimiento, '%Y%m%d').date()
            }
            
        except Exception as e:
            print(f"❌ Error en autorización AFIP: {e}")
            return {
                'success': False,
                'error': str(e),
                'cae': None,
                'vto_cae': None,
                'estado': 'error_afip'
            }

    def get_codigo_iva_afip(self, porcentaje):
        """Mapear porcentajes de IVA a códigos AFIP"""
        mapeo_iva = {
            0: 3,      # Exento
            10.5: 4,   # IVA 10.5%
            21: 5,     # IVA 21%
            27: 6,     # IVA 27%
            2.5: 9     # IVA 2.5%
        }
        
        codigo = mapeo_iva.get(porcentaje, None)
        if codigo is None:
            print(f"⚠️ Porcentaje IVA {porcentaje}% no reconocido, usando código 5 (21%) por defecto")
            return 5  # Por defecto IVA 21%
        
        return codigo
    

    def get_ultimo_comprobante(self, tipo_cbte):
        """Obtener último comprobante autorizado"""
        try:
            print(f"📋 Consultando último comprobante tipo {tipo_cbte}...")
            
            if not self.get_ticket_access():
                raise Exception("No se pudo obtener acceso a AFIP")
            
            # URL del WSFEv1
            wsfe_url = self.config.WSFEv1_URL
            
            print(f"🌐 Conectando con WSFEv1: {wsfe_url}")
            
            # USAR SESIÓN PERSONALIZADA
            session = crear_session_afip()
            
            from zeep.transports import Transport
            
            transport = Transport(session=session, timeout=60)
            client = Client(wsfe_url, transport=transport)
            
            response = client.service.FECompUltimoAutorizado(
                Auth={
                    'Token': self.token,
                    'Sign': self.sign,
                    'Cuit': self.cuit
                },
                PtoVta=self.config.PUNTO_VENTA,
                CbteTipo=tipo_cbte
            )
            
            if hasattr(response, 'Errors') and response.Errors:
                error_msg = response.Errors.Err[0].Msg
                raise Exception(f"Error AFIP: {error_msg}")
            
            ultimo_num = response.CbteNro
            print(f"✅ Último comprobante: {ultimo_num}")
            return ultimo_num
            
        except Exception as e:
            print(f"❌ Error consultando comprobante: {e}")
            raise Exception(f"Error al obtener último comprobante: {e}")


arca_client = ARCAClient()

# Monitor AFIP simplificado
class AFIPStatusMonitor:
    def __init__(self, arca_config):
        self.config = arca_config
    
    def verificar_rapido(self):
        """Verificación rápida solo de conectividad"""
        try:
            import socket
            from urllib.parse import urlparse
            
            wsaa_host = urlparse(self.config.WSAA_URL).hostname
            
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((wsaa_host, 443))
            sock.close()
            
            return {
                'conectividad': result == 0,
                'mensaje': '✅ AFIP accesible' if result == 0 else '❌ AFIP no accesible'
            }
        except Exception as e:
            return {
                'conectividad': False,
                'mensaje': f'❌ Error: {str(e)}'
            }

# Crear instancia del monitor
afip_monitor = AFIPStatusMonitor(ARCA_CONFIG)


# DESPUÉS DE DEFINIR LOS MODELOS Y ANTES DE LAS RUTAS:
# Inicializar y registrar el blueprint de estadísticas
estadisticas_bp = init_estadisticas(db, Factura, DetalleFactura, Producto)
app.register_blueprint(estadisticas_bp)


# Rutas de la aplicación
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        usuario = Usuario.query.filter_by(username=username, activo=True).first()
        
        # Login simple sin encriptación
        if usuario and usuario.password_hash == password:
            session['user_id'] = usuario.id
            session['username'] = usuario.username
            session['nombre'] = usuario.nombre
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

    
@app.route('/api/hora_actual')
def hora_actual():
    from datetime import datetime
    import pytz
    
    # Zona horaria de Argentina
    argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
    ahora = datetime.now(argentina_tz)
    
    return jsonify({
        'success': True,
        'timestamp': ahora.isoformat(),
        'fecha_legible': ahora.strftime('%A, %d de %B de %Y'),
        'hora_legible': ahora.strftime('%H:%M:%S'),
        'zona_horaria': 'America/Argentina/Buenos_Aires'
    })

@app.route('/productos')
def productos():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    productos = Producto.query.filter_by(activo=True).all()
    return render_template('productos.html', productos=productos)

@app.route('/combos')
def combos():
    # Obtener solo productos que son combos
    combos = Producto.query.filter_by(es_combo=True).all()
    return render_template('combos.html', combos=combos)

@app.route('/clientes')
def clientes():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    clientes = Cliente.query.all()
    return render_template('clientes.html', clientes=clientes)

# ==================== RUTAS DE CLIENTES ====================

@app.route('/api/cliente/<int:cliente_id>')
def obtener_cliente(cliente_id):
    """Obtener datos de un cliente por ID para edición"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Método compatible con SQLAlchemy antiguo
        cliente = Cliente.query.get_or_404(cliente_id)
        return jsonify({
            'id': cliente.id,
            'nombre': cliente.nombre,
            'documento': cliente.documento or '',
            'tipo_documento': cliente.tipo_documento or 'DNI',
            'email': cliente.email or '',
            'telefono': cliente.telefono or '',
            'direccion': cliente.direccion or '',
            'condicion_iva': cliente.condicion_iva or 'CONSUMIDOR_FINAL'
        })
    except Exception as e:
        return jsonify({'error': f'Error al obtener cliente: {str(e)}'}), 500

@app.route('/guardar_cliente', methods=['POST'])
def guardar_cliente():
    """Crear o actualizar un cliente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # Validar datos requeridos
        if not data.get('nombre', '').strip():
            return jsonify({'error': 'El nombre es obligatorio'}), 400
        
        # Validar CUIT si es necesario
        if data.get('tipo_documento') == 'CUIT' and data.get('documento'):
            documento = data['documento'].strip()
            if not documento.isdigit() or len(documento) != 11:
                return jsonify({'error': 'El CUIT debe tener 11 dígitos sin guiones'}), 400
        
        cliente_id = data.get('id')
        
        if cliente_id:  # Editar cliente existente
            cliente = Cliente.query.get_or_404(cliente_id)
            accion = 'actualizado'
        else:  # Crear nuevo cliente
            cliente = Cliente()
            accion = 'creado'
        
        # Actualizar datos del cliente
        cliente.nombre = data['nombre'].strip()
        cliente.documento = data.get('documento', '').strip() or None
        cliente.tipo_documento = data.get('tipo_documento', 'DNI')
        cliente.email = data.get('email', '').strip() or None
        cliente.telefono = data.get('telefono', '').strip() or None
        cliente.direccion = data.get('direccion', '').strip() or None
        cliente.condicion_iva = data.get('condicion_iva', 'CONSUMIDOR_FINAL')
        
        # Guardar en base de datos
        if not cliente_id:  # Solo agregar si es nuevo
            db.session.add(cliente)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Cliente {accion} correctamente',
            'cliente_id': cliente.id,
            'cliente_nombre': cliente.nombre
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error guardando cliente: {str(e)}")
        return jsonify({'error': f'Error al guardar cliente: {str(e)}'}), 500

@app.route('/eliminar_cliente/<int:cliente_id>', methods=['DELETE'])
def eliminar_cliente(cliente_id):
    """Eliminar un cliente (marcar como inactivo)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        cliente = Cliente.query.get_or_404(cliente_id)
        
        # Verificar si tiene facturas asociadas
        facturas_count = Factura.query.filter_by(cliente_id=cliente_id).count()
        
        if facturas_count > 0:
            return jsonify({
                'error': f'No se puede eliminar el cliente porque tiene {facturas_count} facturas asociadas'
            }), 400
        
        # Si no tiene facturas, se puede eliminar
        db.session.delete(cliente)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cliente eliminado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error eliminando cliente: {str(e)}")
        return jsonify({'error': f'Error al eliminar cliente: {str(e)}'}), 500

@app.route('/buscar_clientes')
def buscar_clientes():
    """Buscar clientes con filtros"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros de búsqueda
        buscar = request.args.get('buscar', '').strip()
        tipo_doc = request.args.get('tipo_documento', '').strip()
        condicion_iva = request.args.get('condicion_iva', '').strip()
        
        # Construir query base
        query = Cliente.query
        
        # Aplicar filtros
        if buscar:
            query = query.filter(
                or_(
                    Cliente.nombre.ilike(f'%{buscar}%'),
                    Cliente.documento.ilike(f'%{buscar}%'),
                    Cliente.email.ilike(f'%{buscar}%')
                )
            )
        
        if tipo_doc:
            query = query.filter(Cliente.tipo_documento == tipo_doc)
        
        if condicion_iva:
            query = query.filter(Cliente.condicion_iva == condicion_iva)
        
        # Obtener resultados
        clientes = query.order_by(Cliente.nombre).all()
        
        # Formatear respuesta
        resultado = []
        for cliente in clientes:
            resultado.append({
                'id': cliente.id,
                'nombre': cliente.nombre,
                'documento': cliente.documento,
                'tipo_documento': cliente.tipo_documento,
                'email': cliente.email,
                'telefono': cliente.telefono,
                'direccion': cliente.direccion,
                'condicion_iva': cliente.condicion_iva
            })
        
        return jsonify({
            'success': True,
            'clientes': resultado,
            'total': len(resultado)
        })
        
    except Exception as e:
        print(f"Error buscando clientes: {str(e)}")
        return jsonify({'error': f'Error en la búsqueda: {str(e)}'}), 500

# ==================== RUTAS DE PRODUCTOS ====================

# 1. ACTUALIZAR LA RUTA /api/producto_detalle/<int:producto_id>
@app.route('/api/producto_detalle/<int:producto_id>')
def obtener_producto_detalle(producto_id):
    """Obtener datos completos de un producto para edición"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        # Usar valores por defecto si costo o margen son 0 o None
        costo = float(producto.costo) if producto.costo and producto.costo > 0 else 0.0
        margen = float(producto.margen) if producto.margen is not None else 30.0
        
        # Si no hay costo pero hay precio, calcular costo aproximado
        if costo == 0.0 and producto.precio > 0:
            costo = float(producto.precio) / (1 + (margen / 100))
        
        # Datos base del producto
        resultado = {
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'descripcion': producto.descripcion or '',
            'precio': float(producto.precio),
            'costo': round(costo, 2),
            'margen': float(margen),
            'stock': producto.stock,
            'categoria': producto.categoria or '',
            'iva': float(producto.iva),
            'activo': producto.activo,
            
            # ✅ CAMPOS DE COMBO AGREGADOS
            'es_combo': getattr(producto, 'es_combo', False),
            'cantidad_combo': getattr(producto, 'cantidad_combo', None),
            'producto_base_id': getattr(producto, 'producto_base_id', None)
        }
        
        # ✅ SI ES UN COMBO, AGREGAR INFORMACIÓN DEL PRODUCTO BASE
        if resultado['es_combo'] and resultado['producto_base_id']:
            try:
                producto_base = Producto.query.get(resultado['producto_base_id'])
                if producto_base:
                    resultado['producto_base'] = {
                        'id': producto_base.id,
                        'codigo': producto_base.codigo,
                        'nombre': producto_base.nombre,
                        'precio': float(producto_base.precio)
                    }
                    # Precio unitario del producto base para cálculos
                    resultado['precio_unitario_base'] = float(producto_base.precio)
                else:
                    resultado['producto_base'] = None
                    resultado['precio_unitario_base'] = 0.0
            except Exception as e:
                print(f"Error obteniendo producto base: {str(e)}")
                resultado['producto_base'] = None
                resultado['precio_unitario_base'] = 0.0
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"Error en obtener_producto_detalle: {str(e)}")
        return jsonify({'error': f'Error al obtener producto: {str(e)}'}), 500

# 2. ACTUALIZAR LA RUTA /guardar_producto
@app.route('/guardar_producto', methods=['POST'])
def guardar_producto():
    """Crear o actualizar un producto con costo y margen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # Validar datos requeridos
        if not data.get('codigo', '').strip():
            return jsonify({'error': 'El código es obligatorio'}), 400
        
        if not data.get('nombre', '').strip():
            return jsonify({'error': 'El nombre es obligatorio'}), 400
        
        # Validar costo
        try:
            costo = float(data.get('costo', 0))
            if costo <= 0:
                return jsonify({'error': 'El costo debe ser mayor a 0'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Costo inválido'}), 400
        
        # Validar margen
        try:
            margen = float(data.get('margen', 30))
            if margen < 0:
                return jsonify({'error': 'El margen no puede ser negativo'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Margen inválido'}), 400
        
        # Calcular precio automáticamente
        precio_calculado = costo * (1 + (margen / 100))
        
        producto_id = data.get('id')
        codigo = data['codigo'].strip().upper()
        
        # Verificar que el código no exista (excepto si es el mismo producto)
        producto_existente = Producto.query.filter_by(codigo=codigo).first()
        if producto_existente and (not producto_id or producto_existente.id != int(producto_id)):
            return jsonify({'error': f'Ya existe un producto con el código {codigo}'}), 400
        
        if producto_id:  # Editar producto existente
            producto = Producto.query.get_or_404(producto_id)
            accion = 'actualizado'
        else:  # Crear nuevo producto
            producto = Producto()
            accion = 'creado'
        
        # Actualizar datos del producto
        producto.codigo = codigo
        producto.nombre = data['nombre'].strip()
        producto.descripcion = data.get('descripcion', '').strip() or None
        producto.precio = Decimal(str(round(precio_calculado, 2)))
        producto.costo = Decimal(str(round(costo, 2)))
        producto.margen = Decimal(str(round(margen, 2)))
        producto.categoria = data.get('categoria', '').strip() or None
        producto.iva = Decimal(str(data.get('iva', 21)))
        producto.activo = bool(data.get('activo', True))
        producto.fecha_modificacion = datetime.now()
        
        # Solo actualizar stock si es producto nuevo
        if not producto_id:
            producto.stock = int(data.get('stock', 0))
        
        # Guardar en base de datos
        if not producto_id:  # Solo agregar si es nuevo
            db.session.add(producto)
        
        db.session.commit()
        
        print(f"✅ Producto {accion}: {codigo}")
        print(f"   Costo: ${costo:.2f}")
        print(f"   Margen: {margen}%")
        print(f"   Precio: ${precio_calculado:.2f}")
        
        return jsonify({
            'success': True,
            'message': f'Producto {accion} correctamente',
            'producto_id': producto.id,
            'producto_codigo': producto.codigo,
            'precio_calculado': round(precio_calculado, 2),
            'costo': round(costo, 2),
            'margen': round(margen, 2)
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error guardando producto: {str(e)}")
        return jsonify({'error': f'Error al guardar producto: {str(e)}'}), 500




@app.route('/ajustar_stock', methods=['POST'])
def ajustar_stock():
    """Ajustar stock de un producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        producto_id = data.get('producto_id')
        tipo_movimiento = data.get('tipo_movimiento')  # entrada, salida, ajuste
        cantidad = int(data.get('cantidad', 0))
        motivo = data.get('motivo', '').strip()
        
        if not producto_id:
            return jsonify({'error': 'ID de producto requerido'}), 400
        
        if cantidad <= 0 and tipo_movimiento != 'ajuste':
            return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400
        
        producto = Producto.query.get_or_404(producto_id)
        stock_anterior = producto.stock
        
        # Aplicar movimiento según el tipo
        if tipo_movimiento == 'entrada':
            producto.stock += cantidad
            descripcion = f"Entrada: +{cantidad}"
        elif tipo_movimiento == 'salida':
            if cantidad > producto.stock:
                return jsonify({'error': f'No hay suficiente stock. Stock actual: {producto.stock}'}), 400
            producto.stock -= cantidad
            descripcion = f"Salida: -{cantidad}"
        elif tipo_movimiento == 'ajuste':
            if cantidad < 0:
                return jsonify({'error': 'La cantidad para ajuste no puede ser negativa'}), 400
            descripcion = f"Ajuste: {stock_anterior} → {cantidad}"
            producto.stock = cantidad
        else:
            return jsonify({'error': 'Tipo de movimiento inválido'}), 400
        
        # Guardar cambios
        db.session.commit()
        
        # Registrar el movimiento (opcional - puedes crear una tabla de movimientos)
        print(f"MOVIMIENTO STOCK: Producto {producto.codigo} - {descripcion} - Motivo: {motivo}")
        
        return jsonify({
            'success': True,
            'message': f'Stock ajustado correctamente',
            'stock_anterior': stock_anterior,
            'stock_nuevo': producto.stock,
            'movimiento': descripcion
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error ajustando stock: {str(e)}")
        return jsonify({'error': f'Error al ajustar stock: {str(e)}'}), 500

@app.route('/toggle_producto/<int:producto_id>', methods=['POST'])
def toggle_producto(producto_id):
    """Activar/desactivar un producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        # Cambiar estado
        producto.activo = not producto.activo
        estado = 'activado' if producto.activo else 'desactivado'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Producto {estado} correctamente',
            'activo': producto.activo
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error cambiando estado del producto: {str(e)}")
        return jsonify({'error': f'Error al cambiar estado: {str(e)}'}), 500

# 3. ACTUALIZAR LA RUTA /buscar_productos_admin
@app.route('/buscar_productos_admin')
def buscar_productos_admin():
    """Buscar productos con filtros para administración - INCLUYE FILTROS PARA COMBOS"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros de búsqueda
        buscar = request.args.get('buscar', '').strip()
        categoria = request.args.get('categoria', '').strip()
        filtro_stock = request.args.get('stock', '').strip()
        estado = request.args.get('estado', '').strip()
        
        # ✅ NUEVOS PARÁMETROS PARA COMBOS
        solo_combos = request.args.get('solo_combos', '').strip().lower() == 'true'
        estado = request.args.get('estado', '').strip()  # activo, inactivo
        descuento = request.args.get('descuento', '').strip()  # alto, medio, bajo
        
        print(f"🔍 Búsqueda productos admin:")
        print(f"   Buscar: '{buscar}'")
        print(f"   Solo combos: {solo_combos}")
        print(f"   Estado: '{estado}'")
        print(f"   Descuento: '{descuento}'")
        
        # Construir query base
        query = Producto.query
        
        # ✅ FILTRO PARA SOLO COMBOS
        if solo_combos:
            query = query.filter(Producto.es_combo == True)
            print("   Filtro aplicado: Solo combos")
        
        # Aplicar filtros generales
        if buscar:
            query = query.filter(
                or_(
                    Producto.codigo.ilike(f'%{buscar}%'),
                    Producto.nombre.ilike(f'%{buscar}%'),
                    Producto.descripcion.ilike(f'%{buscar}%')
                )
            )
            print(f"   Filtro aplicado: Búsqueda '{buscar}'")
        
        if categoria:
            query = query.filter(Producto.categoria == categoria)
            print(f"   Filtro aplicado: Categoría '{categoria}'")
        
        if filtro_stock == 'bajo':
            query = query.filter(Producto.stock < 10)
        elif filtro_stock == 'sin_stock':
            query = query.filter(Producto.stock <= 0)
        
        # ✅ FILTRO DE ESTADO (ACTIVO/INACTIVO)
        if estado == 'activo':
            query = query.filter(Producto.activo == True)
            print("   Filtro aplicado: Solo activos")
        elif estado == 'inactivo':
            query = query.filter(Producto.activo == False)
            print("   Filtro aplicado: Solo inactivos")
        
        # Obtener resultados SIN filtro de descuento primero
        productos = query.order_by(Producto.codigo).all()
        print(f"   Productos encontrados (antes filtro descuento): {len(productos)}")
        
        # ✅ APLICAR FILTRO DE DESCUENTO DESPUÉS (solo para combos)
        if descuento and solo_combos:
            productos_filtrados = []
            
            for producto in productos:
                if producto.es_combo and producto.producto_base:
                    # Calcular descuento del combo
                    precio_normal = float(producto.producto_base.precio) * float(producto.cantidad_combo)
                    precio_combo = float(producto.precio)
                    descuento_porcentaje = ((precio_normal - precio_combo) / precio_normal) * 100 if precio_normal > 0 else 0
                    
                    # Aplicar filtro según nivel de descuento
                    if descuento == 'alto' and descuento_porcentaje > 30:
                        productos_filtrados.append(producto)
                    elif descuento == 'medio' and 15 <= descuento_porcentaje <= 30:
                        productos_filtrados.append(producto)
                    elif descuento == 'bajo' and descuento_porcentaje < 15:
                        productos_filtrados.append(producto)
                else:
                    # Si no es combo, incluir sin filtro de descuento
                    productos_filtrados.append(producto)
            
            productos = productos_filtrados
            print(f"   Productos después filtro descuento '{descuento}': {len(productos)}")
        
        # Formatear respuesta
        resultado = []
        for producto in productos:
            # Manejar valores por defecto
            costo = float(producto.costo) if producto.costo else 0.0
            margen = float(producto.margen) if producto.margen is not None else 0.0
            
            # Si no hay costo guardado, calcularlo aproximadamente desde precio
            if costo == 0.0 and producto.precio > 0 and margen > 0:
                costo = float(producto.precio) / (1 + (margen / 100))
            
            producto_dict = {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio': float(producto.precio),
                'costo': round(costo, 2),
                'margen': round(margen, 1),
                'stock': producto.stock,
                'categoria': producto.categoria,
                'iva': float(producto.iva),
                'activo': producto.activo,
                'es_combo': producto.es_combo,
                'acceso_rapido': producto.acceso_rapido if hasattr(producto, 'acceso_rapido') else False,
                'orden_acceso_rapido': producto.orden_acceso_rapido if hasattr(producto, 'orden_acceso_rapido') else 0
            }
            
            # ✅ AGREGAR INFORMACIÓN ESPECÍFICA PARA COMBOS
            if producto.es_combo:
                producto_dict.update({
                    'producto_base_id': producto.producto_base_id,
                    'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
                    'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else 0.0
                })
                
                # Información del producto base
                if producto.producto_base:
                    producto_dict['producto_base'] = {
                        'id': producto.producto_base.id,
                        'codigo': producto.producto_base.codigo,
                        'nombre': producto.producto_base.nombre,
                        'precio': float(producto.producto_base.precio)
                    }
            
            resultado.append(producto_dict)
        
        print(f"✅ Búsqueda completada: {len(resultado)} productos")
        
        return jsonify({
            'success': True,
            'productos': resultado,
            'total': len(resultado),
            'filtros_aplicados': {
                'buscar': buscar,
                'solo_combos': solo_combos,
                'estado': estado,
                'descuento': descuento,
                'categoria': categoria,
                'filtro_stock': filtro_stock
            }
        })
        
    except Exception as e:
        print(f"❌ Error buscando productos: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({'error': f'Error en la búsqueda: {str(e)}'}), 500

# FUNCIÓN PARA ACTUALIZAR PRODUCTOS EXISTENTES CON COSTO CALCULADO
@app.route('/actualizar_costos_productos', methods=['POST'])
def actualizar_costos_productos():
    """Actualizar productos que tienen costo 0 calculándolo desde precio y margen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        productos_sin_costo = Producto.query.filter(
            or_(Producto.costo == 0, Producto.costo.is_(None))
        ).all()
        
        contador_actualizados = 0
        
        for producto in productos_sin_costo:
            if producto.precio > 0:
                margen = float(producto.margen) if producto.margen else 30.0
                # Calcular costo desde precio: costo = precio / (1 + margen/100)
                costo_calculado = float(producto.precio) / (1 + (margen / 100))
                
                producto.costo = Decimal(str(round(costo_calculado, 2)))
                producto.fecha_modificacion = datetime.now()
                
                contador_actualizados += 1
                print(f"📦 Actualizado: {producto.codigo} - Precio=${float(producto.precio):.2f} → Costo=${costo_calculado:.2f}")
        
        if contador_actualizados > 0:
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Se actualizaron {contador_actualizados} productos',
            'productos_actualizados': contador_actualizados
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error actualizando costos: {str(e)}")
        return jsonify({'error': f'Error al actualizar costos: {str(e)}'}), 500

@app.route('/obtener_categorias')
def obtener_categorias():
    """Obtener lista de categorías únicas"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener categorías únicas de productos existentes
        categorias = db.session.query(Producto.categoria).filter(
            Producto.categoria.isnot(None),
            Producto.categoria != ''
        ).distinct().all()
        
        categorias_lista = [cat[0] for cat in categorias if cat[0]]
        categorias_lista.sort()
        
        return jsonify({
            'success': True,
            'categorias': categorias_lista
        })
        
    except Exception as e:
        print(f"Error obteniendo categorías: {str(e)}")
        return jsonify({'error': f'Error al obtener categorías: {str(e)}'}), 500


###### RUTAS API PARA MANEJAR COMBOS

@app.route('/api/productos_con_ofertas')
def api_productos_con_ofertas():
    """API para obtener productos con sus ofertas/combos"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        productos = Producto.obtener_productos_con_ofertas()
        return jsonify({
            'success': True,
            'productos': productos
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/crear_combo', methods=['POST'])
def crear_combo():
    """Crear un nuevo combo/oferta basado en un producto existente o editarlo"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # ✅ DETECTAR SI ES EDICIÓN
        combo_id = data.get('id')
        es_edicion = combo_id is not None
        
        if es_edicion:
            print(f"🔧 EDITANDO combo ID: {combo_id}")
            combo = Producto.query.get_or_404(combo_id)
            if not combo.es_combo:
                return jsonify({'error': 'El producto no es un combo'}), 400
        else:
            print("🆕 CREANDO nuevo combo")
            combo = None
        
        # Validar datos
        producto_base_id = data.get('producto_base_id')
        cantidad_combo = float(data.get('cantidad_combo', 1))
        precio_combo = float(data.get('precio_combo', 0))
        nombre_combo = data.get('nombre_combo', '').strip()
        
        if not producto_base_id:
            return jsonify({'error': 'Producto base requerido'}), 400
        
        if cantidad_combo <= 0:
            return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400
        
        if precio_combo <= 0:
            return jsonify({'error': 'El precio debe ser mayor a 0'}), 400
        
        # Obtener producto base
        producto_base = Producto.query.get_or_404(producto_base_id)
        
        # Calcular precio normal y descuento
        precio_normal = float(producto_base.precio) * cantidad_combo
        descuento_monto = precio_normal - precio_combo
        descuento_porcentaje = (descuento_monto / precio_normal) * 100 if precio_normal > 0 else 0
        
        # Generar código automático si no se proporciona
        codigo_combo = data.get('codigo_combo', '').strip()
        if not codigo_combo:
            cantidad_str = f"{cantidad_combo:g}".replace('.', '')
            codigo_combo = f"{producto_base.codigo}-{cantidad_str}UN"
        
        # ✅ VERIFICAR CÓDIGO DUPLICADO SOLO SI ES NUEVO O CAMBIÓ
        if not es_edicion or (es_edicion and combo.codigo != codigo_combo):
            if Producto.query.filter_by(codigo=codigo_combo).first():
                return jsonify({'error': f'Ya existe un producto con el código {codigo_combo}'}), 400
        
        # Generar nombre automático si no se proporciona
        if not nombre_combo:
            cantidad_str = f"{cantidad_combo:g}"
            nombre_combo = f"{cantidad_str} x {producto_base.nombre} (Oferta)"
        
        if es_edicion:
            # ✅ ACTUALIZAR COMBO EXISTENTE
            combo.codigo = codigo_combo
            combo.nombre = nombre_combo
            combo.descripcion = data.get('descripcion_combo', combo.descripcion)
            combo.precio = Decimal(str(precio_combo))
            combo.categoria = 'OFERTAS'
            combo.iva = producto_base.iva
            combo.costo = Decimal(str(float(producto_base.costo or 0) * cantidad_combo))
            combo.stock = int(float(producto_base.stock) / cantidad_combo)
            combo.producto_base_id = producto_base.id
            combo.cantidad_combo = Decimal(str(cantidad_combo))
            combo.precio_unitario_base = producto_base.precio
            combo.descuento_porcentaje = Decimal(str(descuento_porcentaje))
            
            accion = "actualizado"
        else:
            # ✅ CREAR NUEVO COMBO
            combo = Producto(
                codigo=codigo_combo,
                nombre=nombre_combo,
                descripcion=f"Oferta especial: {cantidad_combo:g} unidades de {producto_base.nombre}",
                precio=Decimal(str(precio_combo)),
                categoria='OFERTAS',
                iva=producto_base.iva,
                costo=Decimal(str(float(producto_base.costo or 0) * cantidad_combo)),
                stock=int(float(producto_base.stock) / cantidad_combo),
                es_combo=True,
                producto_base_id=producto_base.id,
                cantidad_combo=Decimal(str(cantidad_combo)),
                precio_unitario_base=producto_base.precio,
                descuento_porcentaje=Decimal(str(descuento_porcentaje))
            )
            
            db.session.add(combo)
            accion = "creado"
        
        db.session.commit()
        
        print(f"✅ Combo {accion}: {codigo_combo}")
        print(f"   Producto base: {producto_base.nombre}")
        print(f"   Cantidad: {cantidad_combo}")
        print(f"   Precio normal: ${precio_normal:.2f}")
        print(f"   Precio combo: ${precio_combo:.2f}")
        print(f"   Descuento: {descuento_porcentaje:.1f}% (${descuento_monto:.2f})")
        
        return jsonify({
            'success': True,
            'message': f'Combo {accion} correctamente',
            'combo': combo.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error en combo: {str(e)}")
        return jsonify({'error': f'Error al procesar combo: {str(e)}'}), 500

@app.route('/api/combos_producto/<int:producto_id>')
def obtener_combos_producto(producto_id):
    """Obtener todos los combos de un producto base"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto_base = Producto.query.get_or_404(producto_id)
        
        # Obtener combos del producto
        combos = Producto.query.filter_by(
            producto_base_id=producto_id,
            es_combo=True,
            activo=True
        ).order_by(Producto.cantidad_combo.asc()).all()
        
        # Preparar respuesta
        combos_data = []
        for combo in combos:
            combo_info = combo.to_dict()
            combos_data.append(combo_info)
        
        return jsonify({
            'success': True,
            'producto_base': producto_base.to_dict(),
            'combos': combos_data,
            'total_combos': len(combos_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# FUNCIÓN PARA MIGRAR PRODUCTOS EXISTENTES
def migrar_productos_para_combos():
    """Migrar productos existentes al nuevo sistema de combos"""
    try:
        print("🔄 Iniciando migración para sistema de combos...")
        
        # Actualizar productos existentes
        productos_sin_migrar = Producto.query.filter(
            Producto.es_combo.is_(None)
        ).all()
        
        contador_migrados = 0
        
        for producto in productos_sin_migrar:
            # Todos los productos existentes son productos base (no combos)
            producto.es_combo = False
            producto.cantidad_combo = Decimal('1.000')
            producto.precio_unitario_base = producto.precio
            producto.descuento_porcentaje = Decimal('0.00')
            
            contador_migrados += 1
        
        if contador_migrados > 0:
            db.session.commit()
            print(f"✅ Migración completada: {contador_migrados} productos actualizados")
        else:
            print("✅ No hay productos para migrar")
            
        return contador_migrados
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en migración: {e}")
        return 0


# EJEMPLO DE USO PARA CREAR COMBOS
def crear_ejemplos_combos():
    """Crear ejemplos de combos para demostración"""
    try:
        # Buscar producto base (milanesa)
        producto_base = Producto.query.filter_by(codigo='MIL001').first()
        
        if not producto_base:
            print("⚠️ Producto base MIL001 no encontrado")
            return
        
        combos_ejemplos = [
            {
                'codigo': 'MIL001-3KG',
                'nombre': '3kg Milanesa de Ternera (Oferta)',
                'cantidad': 3.0,
                'precio': 10000.00,
                'descripcion': 'Oferta especial: 3 kilogramos de milanesa de ternera'
            },
            {
                'codigo': 'MIL001-5KG', 
                'nombre': '5kg Milanesa de Ternera (Super Oferta)',
                'cantidad': 5.0,
                'precio': 18000.00,
                'descripcion': 'Super oferta: 5 kilogramos de milanesa de ternera'
            }
        ]
        
        for combo_data in combos_ejemplos:
            # Verificar si ya existe
            if Producto.query.filter_by(codigo=combo_data['codigo']).first():
                print(f"⚠️ Combo {combo_data['codigo']} ya existe")
                continue
            
            # Calcular valores
            precio_normal = float(producto_base.precio) * combo_data['cantidad']
            descuento_monto = precio_normal - combo_data['precio']
            descuento_porcentaje = (descuento_monto / precio_normal) * 100
            
            # Crear combo
            combo = Producto(
                codigo=combo_data['codigo'],
                nombre=combo_data['nombre'],
                descripcion=combo_data['descripcion'],
                precio=Decimal(str(combo_data['precio'])),
                categoria='OFERTAS',
                iva=producto_base.iva,
                costo=Decimal(str(float(producto_base.costo or 0) * combo_data['cantidad'])),
                stock=int(float(producto_base.stock) / combo_data['cantidad']),
                
                es_combo=True,
                producto_base_id=producto_base.id,
                cantidad_combo=Decimal(str(combo_data['cantidad'])),
                precio_unitario_base=producto_base.precio,
                descuento_porcentaje=Decimal(str(descuento_porcentaje))
            )
            
            db.session.add(combo)
            print(f"✅ Combo creado: {combo_data['codigo']} - Descuento: {descuento_porcentaje:.1f}%")
        
        db.session.commit()
        print("🎉 Ejemplos de combos creados exitosamente")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creando ejemplos: {e}")


# RUTA PARA EJECUTAR MIGRACIÓN
@app.route('/migrar_combos', methods=['POST'])
def ejecutar_migracion_combos():
    """Endpoint para ejecutar migración de combos"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        contador = migrar_productos_para_combos()
        
        return jsonify({
            'success': True,
            'mensaje': f'Migración completada: {contador} productos actualizados',
            'productos_migrados': contador
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en migración: {str(e)}'
        }), 500



@app.route('/nueva_venta')
def nueva_venta():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Obtener productos de acceso rápido en lugar de los primeros 8
    productos = Producto.query.filter_by(
        acceso_rapido=True,
        activo=True
    ).order_by(
        Producto.orden_acceso_rapido.asc(),
        Producto.codigo.asc()
    ).limit(8).all()
    
    clientes = Cliente.query.all()
    return render_template('nueva_venta.html', productos=productos, clientes=clientes)

# APIs para búsqueda de productos
# 4. ACTUALIZAR LAS APIS DE BÚSQUEDA PARA INCLUIR COSTO
@app.route('/api/buscar_productos/<termino>')
def buscar_productos(termino):
    """Busca productos por código o nombre - CORREGIDO"""
    if not termino or len(termino) < 2:
        return jsonify([])
    
    # Búsqueda por código exacto primero
    producto_exacto = Producto.query.filter_by(codigo=termino.upper(), activo=True).first()
    if producto_exacto:
        resultado = {
            'id': producto_exacto.id,
            'codigo': producto_exacto.codigo,
            'nombre': producto_exacto.nombre,
            'precio': float(producto_exacto.precio),
            'precio_base': float(producto_exacto.precio),
            'costo': float(producto_exacto.costo) if producto_exacto.costo else 0.0,
            'margen': float(producto_exacto.margen) if producto_exacto.margen else 0.0,
            'stock': producto_exacto.stock,
            'iva': float(producto_exacto.iva),
            'match_tipo': 'codigo_exacto',
            'descripcion': producto_exacto.descripcion or '',
            'es_combo': producto_exacto.es_combo,
            'producto_base_id': producto_exacto.producto_base_id,
            'cantidad_combo': float(producto_exacto.cantidad_combo) if producto_exacto.cantidad_combo else 1.0,
            'precio_unitario_base': float(producto_exacto.precio_unitario_base) if producto_exacto.precio_unitario_base else float(producto_exacto.precio),
            'descuento_porcentaje': float(producto_exacto.descuento_porcentaje) if producto_exacto.descuento_porcentaje else 0.0,
            'ahorro_combo': producto_exacto.calcular_ahorro_combo(),
            'precio_normal': producto_exacto.calcular_precio_normal(),
            'tiene_ofertas': producto_exacto.tiene_ofertas_volumen()
        }
        return jsonify([resultado])  # ← ESTA LÍNEA ESTABA FALTANDO O MAL UBICADA
    
    # Búsqueda parcial en código y nombre
    termino_busqueda = f"%{termino.lower()}%"
    
    productos = Producto.query.filter(
        and_(
            Producto.activo == True,
            or_(
                Producto.codigo.ilike(termino_busqueda),
                Producto.nombre.ilike(termino_busqueda),
                Producto.descripcion.ilike(termino_busqueda)
            )
        )
    ).limit(15).all()
    
    resultados = []
    for producto in productos:
        # Determinar tipo de coincidencia para ordenar resultados
        match_tipo = 'nombre'
        if termino.lower() in producto.codigo.lower():
            match_tipo = 'codigo'
        elif termino.lower() in producto.nombre.lower()[:20]:
            match_tipo = 'nombre_inicio'
        
        resultado = {
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'precio': float(producto.precio),
            'precio_base': float(producto.precio),
            'costo': float(producto.costo) if producto.costo else 0.0,
            'margen': float(producto.margen) if producto.margen else 0.0,
            'stock': producto.stock,
            'iva': float(producto.iva),
            'match_tipo': match_tipo,
            'descripcion': producto.descripcion or '',
            'es_combo': producto.es_combo,
            'producto_base_id': producto.producto_base_id,
            'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
            'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else float(producto.precio),
            'descuento_porcentaje': float(producto.descuento_porcentaje) if producto.descuento_porcentaje else 0.0,
            'ahorro_combo': producto.calcular_ahorro_combo(),
            'precio_normal': producto.calcular_precio_normal(),
            'tiene_ofertas': producto.tiene_ofertas_volumen()
        }
        resultados.append(resultado)
    
    def orden_relevancia(item):
        if item['match_tipo'] == 'codigo_exacto':
            return 0
        elif item['match_tipo'] == 'codigo':
            return 1
        elif item['match_tipo'] == 'nombre_inicio':
            return 2
        else:
            return 3
    
    resultados.sort(key=orden_relevancia)
    return jsonify(resultados)


@app.route('/api/producto_por_id/<int:producto_id>')
def get_producto_por_id(producto_id):
    """Obtiene un producto por ID - INCLUYE COSTO"""
    producto = Producto.query.filter_by(id=producto_id, activo=True).first()
    if producto:
        return jsonify({
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'precio': float(producto.precio),
            'costo': float(producto.costo) if producto.costo else 0.0,  # ← NUEVO
            'margen': float(producto.margen) if producto.margen else 0.0,  # ← NUEVO
            'stock': producto.stock,
            'iva': float(producto.iva),
            'descripcion': producto.descripcion or '',
            'es_combo': producto.es_combo,
            'producto_base_id': producto.producto_base_id,
            'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
            'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else float(producto.precio),
            'descuento_porcentaje': float(producto.descuento_porcentaje) if producto.descuento_porcentaje else 0.0,
            'ahorro_combo': producto.calcular_ahorro_combo(),
            'precio_normal': producto.calcular_precio_normal()
        })
    return jsonify({'error': 'Producto no encontrado'}), 404

@app.route('/api/producto/<codigo>')
def get_producto(codigo):
    """Obtiene un producto por código exacto - INCLUYE COSTO"""
    producto = Producto.query.filter_by(codigo=codigo.upper(), activo=True).first()
    if producto:
        return jsonify({
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'precio': float(producto.precio),
            'costo': float(producto.costo) if producto.costo else 0.0,  # ← NUEVO
            'margen': float(producto.margen) if producto.margen else 0.0,  # ← NUEVO
            'stock': producto.stock,
            'iva': float(producto.iva),
            'descripcion': producto.descripcion or '',
            'es_combo': producto.es_combo,
            'producto_base_id': producto.producto_base_id,
            'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
            'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else float(producto.precio),
            'descuento_porcentaje': float(producto.descuento_porcentaje) if producto.descuento_porcentaje else 0.0,
            'ahorro_combo': producto.calcular_ahorro_combo(),
            'precio_normal': producto.calcular_precio_normal()

        })
    return jsonify({'error': 'Producto no encontrado'}), 404


# 5. FUNCIÓN AUXILIAR PARA MIGRAR PRODUCTOS EXISTENTES
def migrar_productos_sin_costo_margen():
    """Función para migrar productos existentes que no tienen costo ni margen"""
    try:
        productos_sin_costo = Producto.query.filter(
            or_(
                Producto.costo.is_(None),
                Producto.margen.is_(None),
                Producto.costo == 0
            )
        ).all()
        
        contador_migrados = 0
        
        for producto in productos_sin_costo:
            # Si no tiene costo, calcular desde precio con margen del 30%
            if not producto.costo or producto.costo == 0:
                # Asumiendo un margen del 30%, costo = precio / 1.30
                precio_actual = float(producto.precio)
                costo_calculado = precio_actual / 1.30
                margen_calculado = 30.0
                
                producto.costo = Decimal(str(round(costo_calculado, 2)))
                producto.margen = Decimal(str(margen_calculado))
                
                contador_migrados += 1
                print(f"📦 Migrado: {producto.codigo} - Precio=${precio_actual:.2f} → Costo=${costo_calculado:.2f}, Margen=30%")
        
        if contador_migrados > 0:
            db.session.commit()
            print(f"✅ Migración completada: {contador_migrados} productos actualizados")
        else:
            print("✅ No hay productos que migrar")
            
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en migración: {e}")


# REGISTRAR DESCUENTOEN PROCESAR_VENTA 
def registrar_descuento_factura(factura_id, porcentaje, monto, total_original, usuario_id):
    """Registrar descuento en tabla separada - se llama DESPUÉS de crear factura"""
    try:
        if porcentaje > 0 and monto > 0:
            descuento = DescuentoFactura(
                factura_id=factura_id,
                porcentaje_descuento=Decimal(str(porcentaje)),
                monto_descuento=Decimal(str(monto)),
                total_original=Decimal(str(total_original)),
                usuario_id=usuario_id
            )
            db.session.add(descuento)
            db.session.commit()
            print(f"Descuento registrado: {porcentaje}% = ${monto} para factura {factura_id}")
            return True
    except Exception as e:
        print(f"Error registrando descuento: {e}")
        return False



# FUNCIÓN PROCESAR_VENTA

@app.route('/procesar_venta', methods=['POST'])
def procesar_venta():
    """Procesar venta con medios de pago y items detallados para AFIP"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # Validar datos básicos
        cliente_id = data.get('cliente_id')
        tipo_comprobante = data.get('tipo_comprobante')
        items = data.get('items', [])
        items_detalle = data.get('items_detalle', [])  # *** NUEVO: Items con detalle IVA ***
        medios_pago = data.get('medios_pago', [])
        imprimir_automatico = data.get('imprimir_automatico', True)
        
        if not items:
            return jsonify({'success': False, 'error': 'No hay productos en la venta'})
        
        if not medios_pago:
            return jsonify({'success': False, 'error': 'No se especificaron medios de pago'})
        
        if not items_detalle:
            return jsonify({'success': False, 'error': 'No se recibieron items detallados para AFIP'})
        
        # VALIDACIÓN CORREGIDA PARA VUELTO EN EFECTIVO
        total_medios = sum(float(mp.get('importe', 0)) for mp in medios_pago)
        total_venta = float(data.get('total', 0))
        descuento_monto = float(data.get('descuento_monto', 0))
        descuento_porcentaje = float(data.get('descuento_porcentaje', 0))

        print(f"🔍 DEBUG VALIDACIÓN BACKEND:")
        print(f"   Total medios de pago: ${total_medios:.2f}")
        print(f"   Total venta (con descuento): ${total_venta:.2f}")
        print(f"   Descuento aplicado: ${descuento_monto:.2f}")

        diferencia = total_medios - total_venta

        if diferencia < -0.01:
            # Falta dinero - ERROR
            faltante = abs(diferencia)
            print(f"❌ ERROR: Faltan ${faltante:.2f}")
            print(f"   Medios individuales:")
            for i, mp in enumerate(medios_pago):
                print(f"     {i+1}. {mp.get('medio_pago', 'UNKNOWN')}: ${float(mp.get('importe', 0)):.2f}")
            
            return jsonify({
                'success': False, 
                'error': f'Faltan ${faltante:.2f} para completar el pago'
            })
        elif diferencia > 0.01:
            # Hay exceso - verificar si es efectivo para dar vuelto
            medios_efectivo = [mp for mp in medios_pago if mp.get('medio_pago') == 'efectivo']
            total_efectivo = sum(float(mp.get('importe', 0)) for mp in medios_efectivo)
            
            # Si todo el exceso puede cubrirse con efectivo, está OK
            if total_efectivo >= total_venta:
                vuelto = diferencia
                print(f"✅ Pago con vuelto: ${vuelto:.2f}")
                # Continuar procesando - el vuelto se maneja en el frontend
            else:
                # Hay exceso pero no suficiente efectivo para vuelto
                print(f"❌ ERROR: Exceso de ${diferencia:.2f} sin suficiente efectivo")
                return jsonify({
                    'success': False,
                    'error': f'Exceso de ${diferencia:.2f} pero no hay suficiente efectivo para dar vuelto'
                })
        else:
            print(f"✅ Pago exacto")

        print(f"✅ Validación exitosa. Procediendo con la venta.")
        
        # PASO 1: Determinar el próximo número de comprobante
        tipo_comprobante_int = int(tipo_comprobante)
        punto_venta = ARCA_CONFIG.PUNTO_VENTA
        
        # PASO 2: Generar número temporal primero
        ultima_factura_local = Factura.query.filter_by(
            tipo_comprobante=str(tipo_comprobante_int),
            punto_venta=punto_venta
        ).order_by(Factura.id.desc()).first()
        
        numero_temporal = 1
        if ultima_factura_local and ultima_factura_local.numero:
            try:
                ultimo_numero_local = int(ultima_factura_local.numero.split('-')[1])
                numero_temporal = ultimo_numero_local + 1
            except:
                numero_temporal = 1
        
        # Generar número temporal para la base de datos
        numero_factura_temporal = f"{punto_venta:04d}-{numero_temporal:08d}"
        
        # Verificar que el número temporal no exista
        factura_existente = Factura.query.filter_by(numero=numero_factura_temporal).first()
        while factura_existente:
            numero_temporal += 1
            numero_factura_temporal = f"{punto_venta:04d}-{numero_temporal:08d}"
            factura_existente = Factura.query.filter_by(numero=numero_factura_temporal).first()
        
        print(f"📝 Número temporal asignado: {numero_factura_temporal}")
        
        # PASO 3: Crear factura CON número temporal
        total_final = float(data.get('total', total_venta))  # Usar el total que ya tiene descuento
        factura = Factura(
            numero=numero_factura_temporal,
            tipo_comprobante=str(tipo_comprobante_int),
            punto_venta=punto_venta,
            cliente_id=cliente_id,
            usuario_id=session['user_id'],
            subtotal=Decimal(str(data['subtotal'])),
            iva=Decimal(str(data['iva'])),
            total=Decimal(str(total_venta))
        )
        
        db.session.add(factura)
        db.session.flush()  # Para obtener el ID sin hacer commit
        
        print(f"✅ Factura creada con ID: {factura.id} y número temporal: {factura.numero}")
        
       # PASO 4: Agregar detalles de productos CON IVA CORRECTO
        for i, item in enumerate(items):
            # Obtener el detalle correspondiente con IVA
            item_detalle = items_detalle[i] if i < len(items_detalle) else {}
            iva_porcentaje = float(item_detalle.get('iva_porcentaje', 21.0))
            
            # Calcular IVA del item con redondeo AFIP
            subtotal = float(item['subtotal'])
            importe_iva = round((subtotal * iva_porcentaje / 100), 2)
            
            detalle = DetalleFactura(
                factura_id=factura.id,
                producto_id=item['producto_id'],
                cantidad=item['cantidad'],
                precio_unitario=Decimal(str(item['precio_unitario'])),
                subtotal=Decimal(str(subtotal)),
                porcentaje_iva=Decimal(str(iva_porcentaje)),  # ✅ GUARDAR IVA CORRECTO
                importe_iva=Decimal(str(importe_iva))          # ✅ GUARDAR IMPORTE IVA
            )
            db.session.add(detalle)
            
            # Actualizar stock
            producto = Producto.query.get(item['producto_id'])
            if producto:
                producto.stock -= Decimal(str(item['cantidad']))
                print(f"📦 Stock actualizado para {producto.nombre}: {producto.stock}")
                print(f"💰 Detalle guardado: IVA {iva_porcentaje}% = ${importe_iva:.2f}")
            else:
                print(f"⚠️ Producto ID {item['producto_id']} no encontrado")
        
        # PASO 5: Agregar medios de pago
        print(f"💳 Agregando {len(medios_pago)} medios de pago...")
        for medio_data in medios_pago:
            medio_pago = MedioPago(
                factura_id=factura.id,
                medio_pago=medio_data['medio_pago'],
                importe=Decimal(str(medio_data['importe'])),
                fecha_registro=datetime.now()
            )
            db.session.add(medio_pago)
            print(f"💰 Medio agregado: {medio_data['medio_pago']} ${medio_data['importe']}")
        
        # PASO 6: Intentar autorizar en AFIP con items detallados
        try:
            print("📄 Autorizando en AFIP con items detallados...")
            cliente = Cliente.query.get(cliente_id)
            
            # *** CLAVE: Preparar datos para AFIP con items_detalle ***
            datos_comprobante = {
                'tipo_comprobante': tipo_comprobante_int,
                'punto_venta': punto_venta,
                'importe_neto': float(factura.subtotal),
                'importe_iva': float(factura.iva),
                'items_detalle': items_detalle,  # *** NUEVO: Items con IVA individual ***
                'doc_tipo': 99,  # Sin identificar por defecto
                'doc_nro': 0
            }
            
            # Debug: Mostrar items que se envían a AFIP
            print("🧮 Items detallados enviados a AFIP:")
            for item in items_detalle:
                print(f"   📦 ${item['subtotal']:.2f} (IVA {item['iva_porcentaje']}%)")
            
            # Si el cliente tiene documento, agregar datos
            if cliente and cliente.documento:
                if cliente.tipo_documento == 'CUIT' and len(cliente.documento) == 11:
                    datos_comprobante['doc_tipo'] = 80  # CUIT
                    datos_comprobante['doc_nro'] = int(cliente.documento)
                elif cliente.tipo_documento == 'DNI' and len(cliente.documento) >= 7:
                    datos_comprobante['doc_tipo'] = 96  # DNI
                    datos_comprobante['doc_nro'] = int(cliente.documento)
            
            resultado_afip = arca_client.autorizar_comprobante(datos_comprobante)
            
            if resultado_afip['success']:
                # AFIP devuelve el número correcto - ACTUALIZAR la factura
                numero_afip = resultado_afip['numero']
                print(f"✅ AFIP asignó número: {numero_afip}")
                
                # Verificar si el número de AFIP ya existe en la BD
                factura_afip_existente = Factura.query.filter(
                    and_(Factura.numero == numero_afip, Factura.id != factura.id)
                ).first()
                
                if factura_afip_existente:
                    print(f"⚠️ Número AFIP {numero_afip} ya existe, manteniendo temporal")
                    # Mantener número temporal y marcar como autorizada
                    factura.cae = resultado_afip['cae']
                    factura.vto_cae = resultado_afip['vto_cae']
                    factura.estado = 'autorizada'
                else:
                    # Actualizar con número de AFIP
                    factura.numero = numero_afip
                    factura.cae = resultado_afip['cae']
                    factura.vto_cae = resultado_afip['vto_cae']
                    factura.estado = 'autorizada'
                
                print(f"✅ Autorización AFIP exitosa. CAE: {factura.cae}")
                print(f"✅ Número final: {factura.numero}")
            else:
                # Si AFIP falla, mantener número temporal
                factura.estado = 'error_afip'
                print(f"❌ Error AFIP: {resultado_afip.get('error', 'Error desconocido')}")
                print(f"📝 Manteniendo número temporal: {factura.numero}")
            
        except Exception as e:
            # Si hay error completo, mantener número temporal
            factura.estado = 'error_afip'
            print(f"❌ Error completo al autorizar en AFIP: {e}")
            print(f"📝 Manteniendo número temporal: {factura.numero}")
            
        # PASO 7: Guardar todo en la base de datos
        db.session.commit()
        
        print(f"🎉 Venta procesada exitosamente: {factura.numero}")
        
        # NUEVA LÍNEA: Registrar descuento si existe (tabla separada)
        if data.get('descuento_monto', 0) > 0:
            total_antes_descuento = float(data.get('subtotal', 0)) + float(data.get('iva', 0))
            registrar_descuento_factura(
                factura.id, 
                data.get('descuento_porcentaje', 0),
                data.get('descuento_monto', 0),
                total_antes_descuento,
                session['user_id']
            )

        # PASO 8: Imprimir automáticamente si está configurado
        if imprimir_automatico and IMPRESION_DISPONIBLE:
            try:
                print("🖨️ Imprimiendo factura automáticamente...")
                impresora_termica.imprimir_factura(factura)
            except Exception as e:
                print(f"⚠️ Error en impresión automática: {e}")
        
        return jsonify({
            'success': True,
            'factura_id': factura.id,
            'numero': factura.numero,
            'cae': factura.cae,
            'estado': factura.estado,
            'mensaje': f"Factura {factura.numero} generada correctamente"
        })
        
    except Exception as e:
        print(f"❌ Error en procesar_venta: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Error al procesar la venta: {str(e)}'}), 500


# RUTAS DE IMPRESIÓN
@app.route('/imprimir_factura/<int:factura_id>')
def imprimir_factura(factura_id):
    """Endpoint para imprimir factura específica"""
    try:
        print(f"🖨️ Iniciando impresión de factura ID: {factura_id}")
        
        # Obtener la factura de la base de datos
        factura = Factura.query.get_or_404(factura_id)
        
        # Usar directamente la instancia de la clase
        resultado = impresora_termica.imprimir_factura(factura)
        
        if resultado:
            return jsonify({
                'success': True,
                'mensaje': f'Factura impresa correctamente en {impresora_termica.nombre_impresora}'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Error al imprimir factura'
            })
        
    except Exception as e:
        print(f"❌ Error en endpoint imprimir: {e}")
        return jsonify({
            'success': False,
            'error': f'Error al imprimir factura: {str(e)}'
        }), 500
        
@app.route('/test_impresion')
def test_impresion():
    try:
        print("🧪 Iniciando test de impresión...")
        resultado = impresora_termica.test_impresion()  # ← Llama al método de la clase
        
        if resultado:
            return jsonify({
                'success': True,
                'mensaje': f'Test enviado correctamente a {impresora_termica.nombre_impresora}'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Error al enviar test de impresión'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en test: {str(e)}'
        }), 500

@app.route('/estado_impresora')
def estado_impresora():
    """Endpoint para verificar estado de impresora"""
    try:
        estado = impresora_termica.verificar_estado()
        return jsonify(estado)
    except Exception as e:
        return jsonify({
            'disponible': False,
            'error': f'Error verificando estado: {str(e)}'
        }), 500

@app.route('/facturas')
def facturas():
    """Página de gestión de facturas (carga dinámica via JavaScript)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # No cargar facturas aquí - se cargan dinámicamente via JavaScript
    # Esto mejora significativamente el tiempo de carga de la página
    
    return render_template('facturas.html', facturas=[])

@app.route('/factura/<int:factura_id>')
def ver_factura(factura_id):
    try:
        # Método compatible con SQLAlchemy antiguo
        factura = Factura.query.get_or_404(factura_id)
        
        return render_template('factura_detalle.html', factura=factura)
        
    except Exception as e:
        app.logger.error(f"Error al cargar factura {factura_id}: {str(e)}")
        flash('Error al cargar la factura', 'error')
        return redirect(url_for('facturas'))

# Funciones de utilidad para limpieza de datos
def limpiar_facturas_duplicadas():
    """Limpiar facturas duplicadas o problemáticas"""
    try:
        # Buscar facturas con números duplicados
        facturas_duplicadas = db.session.query(Factura.numero).group_by(Factura.numero).having(db.func.count(Factura.numero) > 1).all()
        
        if facturas_duplicadas:
            print(f"⚠️ Encontradas {len(facturas_duplicadas)} facturas con números duplicados")
            
            for numero_duplicado in facturas_duplicadas:
                numero = numero_duplicado[0]
                facturas = Factura.query.filter_by(numero=numero).order_by(Factura.id).all()
                
                # Mantener solo la primera, eliminar las demás
                for i, factura in enumerate(facturas):
                    if i > 0:  # Eliminar todas excepto la primera
                        print(f"🗑️ Eliminando factura duplicada: {factura.numero} (ID: {factura.id})")
                        
                        # Eliminar detalles primero
                        DetalleFactura.query.filter_by(factura_id=factura.id).delete()
                        
                        # Eliminar factura
                        db.session.delete(factura)
            
            db.session.commit()
            print("✅ Limpieza de duplicados completada")
        else:
            print("✅ No se encontraron facturas duplicadas")
            
    except Exception as e:
        print(f"❌ Error en limpieza: {e}")
        db.session.rollback()

def verificar_estado_facturas():
    """Verificar el estado actual de las facturas"""
    try:
        total_facturas = Factura.query.count()
        facturas_con_cae = Factura.query.filter(Factura.cae.isnot(None)).count()
        facturas_pendientes = Factura.query.filter_by(estado='pendiente').count()
        facturas_error = Factura.query.filter_by(estado='error_afip').count()
        
        print("\n📊 ESTADO ACTUAL DE FACTURAS:")
        print(f"   Total facturas: {total_facturas}")
        print(f"   Con CAE: {facturas_con_cae}")
        print(f"   Pendientes: {facturas_pendientes}")
        print(f"   Con errores: {facturas_error}")
        
        # Mostrar últimas facturas
        ultimas_facturas = Factura.query.order_by(Factura.id.desc()).limit(5).all()
        print(f"\n📋 ÚLTIMAS 5 FACTURAS:")
        for factura in ultimas_facturas:
            cae_status = f"CAE: {factura.cae[:10]}..." if factura.cae else "Sin CAE"
            print(f"   {factura.numero} - {factura.estado} - {cae_status}")
        
    except Exception as e:
        print(f"❌ Error verificando estado: {e}")

# Funciones de inicialización
def create_tables():
    """Crea las tablas de la base de datos"""
    try:
        db.create_all()
        
        # Crear usuario admin por defecto si no existe
        if not Usuario.query.filter_by(username='admin').first():
            admin = Usuario(
                username='admin',
                password_hash='admin123',  # Sin encriptación para simplicidad
                nombre='Administrador',
                rol='admin'
            )
            db.session.add(admin)
            db.session.commit()
            print("✅ Usuario admin creado (admin/admin123)")
        
        print("✅ Base de datos inicializada correctamente")
        
    except Exception as e:
        print(f"❌ Error al inicializar base de datos: {e}")

@app.route('/qr_afip/<int:factura_id>')
def generar_qr_afip(factura_id):
    try:
        # Método compatible con SQLAlchemy antiguo
        factura = Factura.query.get_or_404(factura_id)
        
        try:
            generador_qr = crear_generador_qr(ARCA_CONFIG)  # ← Ya está correcto
            
            # Generar imagen QR
            qr_base64 = generador_qr.generar_qr_imagen(factura)
            info_qr = generador_qr.obtener_info_qr(factura)
            
            if qr_base64:
                return jsonify({
                    'success': True,
                    'qr_image': qr_base64,
                    'qr_url': info_qr['url'],
                    'qr_valido': info_qr['valido'],
                    'mensaje': info_qr['mensaje']
                })
            else:
                return jsonify({
                    'success': False,
                    'error': 'Error al generar código QR'
                }), 500
                
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'Módulo QR no disponible'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error: {str(e)}'
        }), 500

# Ruta para mostrar QR en página completa (para escanear con móvil)
@app.route('/mostrar_qr/<int:factura_id>')
def mostrar_qr_completo(factura_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Método compatible con SQLAlchemy antiguo
    factura = Factura.query.get_or_404(factura_id)
    
    try:
        generador_qr = crear_generador_qr(ARCA_CONFIG)  # ← Ya está correcto
        info_qr = generador_qr.obtener_info_qr(factura)
        qr_base64 = generador_qr.generar_qr_imagen(factura, tamaño=8) if info_qr['valido'] else None
    except Exception as e:
        print(f"⚠️ Error generando QR: {e}")
        info_qr = {'valido': False, 'mensaje': f'Error al generar QR: {str(e)}'}
        qr_base64 = None
    
    return render_template('mostrar_qr.html', 
                         factura=factura, 
                         qr_info=info_qr,
                         qr_image=qr_base64)

# Ruta para validar datos QR
@app.route('/validar_qr/<int:factura_id>')
def validar_qr_factura(factura_id):
    try:
        # Método compatible con SQLAlchemy antiguo
        factura = Factura.query.get_or_404(factura_id)
        
        try:
            generador_qr = crear_generador_qr(ARCA_CONFIG)  # ← Ya está correcto
            info_qr = generador_qr.obtener_info_qr(factura)
            errores = generador_qr.validar_datos_qr(factura)
        except Exception as e:
            print(f"⚠️ Error con módulo QR: {e}")
            info_qr = {'valido': False, 'mensaje': f'Error con módulo QR: {str(e)}'}
            errores = [str(e)]
        
        return jsonify({
            'factura_id': factura_id,
            'numero': factura.numero,
            'qr_valido': info_qr['valido'],
            'qr_url': info_qr['url'] if info_qr['valido'] else None,
            'errores': errores,
            'datos_qr': info_qr.get('datos', {}),
            'mensaje': info_qr['mensaje']
        })
        
    except Exception as e:
        return jsonify({
            'error': f'Error validando QR: {str(e)}'
        }), 500



@app.route('/api/estado_afip_rapido')
def api_estado_afip_rapido():
    """API para verificación rápida de AFIP"""
    try:
        estado = afip_monitor.verificar_rapido()
        return jsonify({
            'success': True,
            'estado': estado
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/test_afip')
def test_afip():
    """Test manual de conexión AFIP"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        resultado_auth = arca_client.get_ticket_access()
        
        return jsonify({
            'success': resultado_auth,
            'mensaje': "✅ Test AFIP exitoso" if resultado_auth else "❌ Test AFIP falló",
            'estado': "success" if resultado_auth else "error",
            'tiene_token': bool(arca_client.token),
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'mensaje': f"❌ Error en test: {str(e)}",
            'estado': 'error'
        }), 500


# Agregar esta función en app.py para debug de WSFEv1

# Versión simplificada sin verificación de sesión para debug
##### http://localhost:5000/debug_afip_simple LLAMAR ESTA RUTA PARA VERIFICAR ARCA

@app.route('/debug_afip_simple')
def debug_afip_simple():
    """Debug simple de AFIP sin verificación de sesión"""
    try:
        resultado = {
            'timestamp': datetime.now().isoformat(),
            'tests': {}
        }
        
        print("🔍 INICIANDO DEBUG SIMPLE DE AFIP...")
        
        # Test 1: Autenticación WSAA
        print("1️⃣ Test autenticación WSAA...")
        try:
            auth_result = arca_client.get_ticket_access()
            resultado['tests']['wsaa_auth'] = {
                'success': auth_result,
                'token_exists': bool(arca_client.token),
                'message': 'Autenticación exitosa' if auth_result else 'Fallo en autenticación'
            }
        except Exception as e:
            resultado['tests']['wsaa_auth'] = {
                'success': False,
                'error': str(e),
                'message': 'Error en autenticación'
            }
        
        # Test 2: Conectividad WSFEv1
        print("2️⃣ Test conectividad WSFEv1...")
        try:
            import socket
            from urllib.parse import urlparse
            
            wsfe_url = ARCA_CONFIG.WSFEv1_URL
            parsed = urlparse(wsfe_url)
            host = parsed.hostname
            port = 443
            
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(10)
            connect_result = sock.connect_ex((host, port))
            sock.close()
            
            resultado['tests']['wsfe_connectivity'] = {
                'success': connect_result == 0,
                'host': host,
                'port': port,
                'url': wsfe_url,
                'message': 'Conectividad OK' if connect_result == 0 else f'No se puede conectar (código: {connect_result})'
            }
        except Exception as e:
            resultado['tests']['wsfe_connectivity'] = {
                'success': False,
                'error': str(e),
                'message': 'Error verificando conectividad'
            }
        
        # Test 3: Solo si tenemos autenticación, probar FEDummy
        if resultado['tests']['wsaa_auth']['success']:
            print("3️⃣ Test FEDummy...")
            try:
                from zeep import Client
                from zeep.transports import Transport
                
                session_afip = crear_session_afip()
                transport = Transport(session=session_afip, timeout=30)
                client = Client(ARCA_CONFIG.WSFEv1_URL, transport=transport)
                
                dummy_response = client.service.FEDummy()
                
                resultado['tests']['fe_dummy'] = {
                    'success': True,
                    'response': str(dummy_response),
                    'message': 'FEDummy exitoso'
                }
            except Exception as e:
                resultado['tests']['fe_dummy'] = {
                    'success': False,
                    'error': str(e),
                    'message': 'Error en FEDummy'
                }
        
        # Generar resumen
        total_tests = len(resultado['tests'])
        successful_tests = sum(1 for test in resultado['tests'].values() if test['success'])
        
        resultado['summary'] = {
            'total_tests': total_tests,
            'successful_tests': successful_tests,
            'success_rate': f"{(successful_tests/total_tests*100):.1f}%" if total_tests > 0 else "0%",
            'overall_status': 'OK' if successful_tests == total_tests else 'PROBLEMAS'
        }
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error general: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500



@app.route('/reporte_medios_hoy')
def reporte_medios_hoy():
    """Reporte rápido de medios de pago del día actual"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        fin_hoy = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        
        reporte = MedioPago.calcular_recaudacion_por_fecha(hoy, fin_hoy)
        
        return jsonify({
            'success': True,
            'reporte': reporte
        })
        
    except Exception as e:
        print(f"Error en reporte del día: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/medios_pago_factura/<int:factura_id>')
def medios_pago_factura(factura_id):
    """Obtener los medios de pago de una factura específica"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        medios = [{
            'medio_pago': mp.medio_pago,
            'importe': float(mp.importe),
            'fecha_registro': mp.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
        } for mp in factura.medios_pago]
        
        return jsonify({
            'success': True,
            'factura_numero': factura.numero,
            'total_factura': float(factura.total),
            'medios_pago': medios
        })
        
    except Exception as e:
        print(f"Error obteniendo medios de pago: {e}")
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    # Crear directorios necesarios
    os.makedirs('cache', exist_ok=True)
    os.makedirs('logs', exist_ok=True)
    os.makedirs('certificados', exist_ok=True)
    
    print("🚀 Iniciando POS Argentina...")
    print(f"📍 URL: http://localhost:5000")
    print(f"🏢 CUIT: {ARCA_CONFIG.CUIT}")
    print(f"🏪 Punto de Venta: {ARCA_CONFIG.PUNTO_VENTA}")
    print(f"🔧 Ambiente: {'HOMOLOGACIÓN' if ARCA_CONFIG.USE_HOMOLOGACION else 'PRODUCCIÓN'}")
    print(f"🖨️ Impresión: {'Disponible' if IMPRESION_DISPONIBLE else 'No disponible'}")
    print(f"👤 Usuario: admin")
    print(f"🔑 Contraseña: admin123")
    print()
    
    with app.app_context():
        create_tables()
        
        migrar_productos_sin_costo_margen()  # ← EJECUTAR UNA SOLA VEZ

        # Limpiar datos problemáticos
        print("🧹 Verificando integridad de datos...")
        limpiar_facturas_duplicadas()
        verificar_estado_facturas()
    
# AGREGAR esta función a tu app.py (después de la línea que dice @app.route('/medios_pago_factura/<int:factura_id>')):

@app.route('/api/reporte_medios_pago')
def reporte_medios_pago():
    """Generar reporte de medios de pago por fecha usando SQLAlchemy"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros de fecha
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')
        
        if not fecha_desde or not fecha_hasta:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar fechas desde y hasta'
            })
        
        # Convertir strings a datetime
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
            })
        
        # Consulta usando SQLAlchemy
        from sqlalchemy import func
        
        resultados = db.session.query(
            MedioPago.medio_pago,
            func.count(MedioPago.id).label('cantidad'),
            func.sum(MedioPago.importe).label('total')
        ).join(
            Factura, MedioPago.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt
            )
        ).group_by(MedioPago.medio_pago).order_by(
            func.sum(MedioPago.importe).desc()
        ).all()
        
        # Formatear resultados
        medios_pago = []
        total_general = 0
        
        for medio, cantidad, total in resultados:
            medios_pago.append({
                'medio_pago': medio,
                'cantidad': cantidad,
                'total': float(total)
            })
            total_general += float(total)
        
        return jsonify({
            'success': True,
            'reporte': {
                'medios_pago': medios_pago,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'total_general': total_general
            }
        })
        
    except Exception as e:
        print(f"Error en reporte_medios_pago: {e}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/reporte_ventas')
def reporte_ventas():
    """Página principal del reporte de ventas"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('reporte_ventas.html')


@app.route('/api/reporte_ventas_productos')
def api_reporte_ventas_productos():
    """API para generar reporte de ventas por producto - CORREGIDO PARA COMBOS"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        categoria = request.args.get('categoria', '').strip()
        orden = request.args.get('orden', 'cantidad_desc')
        solo_con_ventas = request.args.get('solo_con_ventas', 'true').lower() == 'true'
        
        # Validar fechas
        if not fecha_desde or not fecha_hasta:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar fechas desde y hasta'
            })
        
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido'
            })
        
        # Validar rango de fechas (máximo 2 años)
        if (fecha_hasta_dt - fecha_desde_dt).days > 730:
            return jsonify({
                'success': False,
                'error': 'El rango de fechas no puede ser mayor a 2 años'
            })
        
        print(f"📊 Generando reporte de ventas (CORREGIDO PARA COMBOS):")
        print(f"   Período: {fecha_desde} a {fecha_hasta}")
        print(f"   Categoría: {categoria or 'Todas'}")
        print(f"   Orden: {orden}")
        print(f"   Solo con ventas: {solo_con_ventas}")
        
        # ✅ QUERY CORREGIDA: Incluir cantidad_combo en el cálculo
        query = db.session.query(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo,
            # *** CLAVE: Multiplicar cantidad por cantidad_combo si es combo ***
            func.sum(
                case(
                    (Producto.es_combo == True, DetalleFactura.cantidad * Producto.cantidad_combo),
                    else_=DetalleFactura.cantidad
                )
            ).label('cantidad_real_vendida'),
            func.sum(DetalleFactura.cantidad).label('unidades_combos_vendidas'),  # Para mostrar también combos vendidos
            func.sum(DetalleFactura.subtotal).label('total_vendido'),
            func.avg(DetalleFactura.precio_unitario).label('precio_promedio'),
            func.max(Factura.fecha).label('ultima_venta'),
            func.count(DetalleFactura.id).label('num_transacciones')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt
                # Incluye TODAS las facturas (sin filtro de estado)
            )
        )
        
        print(f"✅ Query configurada con corrección para combos")
        
        # Filtrar por categoría si se especifica
        if categoria:
            query = query.filter(Producto.categoria == categoria)
            print(f"   Filtro categoría aplicado: {categoria}")
        
        # Agrupar por producto
        query = query.group_by(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo
        )
        
        # *** APLICAR ORDENAMIENTO USANDO LA CANTIDAD REAL ***
        if orden == 'cantidad_desc':
            query = query.order_by(desc('cantidad_real_vendida'))
        elif orden == 'cantidad_asc':
            query = query.order_by(asc('cantidad_real_vendida'))
        elif orden == 'total_desc':
            query = query.order_by(desc('total_vendido'))
        elif orden == 'total_asc':
            query = query.order_by(asc('total_vendido'))
        elif orden == 'codigo':
            query = query.order_by(Producto.codigo)
        elif orden == 'nombre':
            query = query.order_by(Producto.nombre)
        
        print(f"   Ordenamiento aplicado: {orden} (usando cantidad real)")
        
        # Ejecutar query
        print(f"🔍 Ejecutando consulta...")
        resultados = query.all()
        print(f"📋 Encontrados {len(resultados)} productos con ventas")
        
        # Consulta adicional: información de estados de facturas para debug
        debug_estados = db.session.query(
            Factura.estado,
            func.count(Factura.id).label('cantidad'),
            func.sum(Factura.total).label('total')
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt
            )
        ).group_by(Factura.estado).all()
        
        estados_info = {}
        for estado, cantidad, total in debug_estados:
            estados_info[estado] = {
                'cantidad': cantidad,
                'total': float(total) if total else 0.0
            }
        
        print(f"📊 Estados de facturas en el período:")
        for estado, info in estados_info.items():
            print(f"   {estado}: {info['cantidad']} facturas (${info['total']:.2f})")
        
        # *** FORMATEAR RESULTADOS CON CANTIDAD REAL ***
        productos = []
        total_unidades_reales = 0
        total_ventas = 0.0
        
        for resultado in resultados:
            # *** USAR CANTIDAD REAL (ya calculada en SQL) ***
            cantidad_real = float(resultado.cantidad_real_vendida) if resultado.cantidad_real_vendida else 0.0
            unidades_combos = int(resultado.unidades_combos_vendidas) if resultado.unidades_combos_vendidas else 0
            total_producto = float(resultado.total_vendido) if resultado.total_vendido else 0.0
            precio_promedio = float(resultado.precio_promedio) if resultado.precio_promedio else 0.0
            
            # *** INFORMACIÓN ADICIONAL PARA COMBOS ***
            info_combo = ""
            unidad_medida = "unidades"
            
            if resultado.es_combo and resultado.cantidad_combo:
                cantidad_combo = float(resultado.cantidad_combo)
                info_combo = f" ({unidades_combos} combos de {cantidad_combo:g} c/u)"
                # Detectar unidad de medida basada en cantidad del combo
                if cantidad_combo >= 1:
                    if cantidad_combo == int(cantidad_combo):
                        unidad_medida = "kg" if cantidad_combo >= 1 else "unidades"
                    else:
                        unidad_medida = "kg"
                
                print(f"📦 {resultado.codigo}: {unidades_combos} combos × {cantidad_combo:g} = {cantidad_real:g} {unidad_medida}")
            
            productos.append({
                'id': resultado.id,
                'codigo': resultado.codigo,
                'nombre': resultado.nombre,
                'descripcion': resultado.descripcion,
                'categoria': resultado.categoria,
                'es_combo': resultado.es_combo,
                'cantidad_combo': float(resultado.cantidad_combo) if resultado.cantidad_combo else 1.0,
                'cantidad_vendida': cantidad_real,  # *** CANTIDAD REAL ***
                'unidades_combos_vendidas': unidades_combos,  # *** COMBOS VENDIDOS ***
                'info_combo': info_combo,  # *** INFORMACIÓN ADICIONAL ***
                'unidad_medida': unidad_medida,  # *** UNIDAD DE MEDIDA ***
                'total_vendido': total_producto,
                'precio_promedio': precio_promedio,
                'ultima_venta': resultado.ultima_venta.isoformat() if resultado.ultima_venta else None,
                'num_transacciones': int(resultado.num_transacciones) if resultado.num_transacciones else 0
            })
            
            total_unidades_reales += cantidad_real
            total_ventas += total_producto
        
        # *** RESUMEN CORREGIDO ***
        resumen = {
            'total_productos': len(productos),
            'total_unidades_reales': total_unidades_reales,  # *** UNIDADES REALES ***
            'total_ventas': total_ventas,
            'promedio_por_producto': total_ventas / len(productos) if len(productos) > 0 else 0,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'categoria_filtro': categoria or 'Todas',
            'incluye_todas_facturas': True,
            'estados_facturas': estados_info,
            'correccion_combos': True  # *** FLAG PARA INDICAR CORRECCIÓN ***
        }
        
        print(f"✅ Reporte generado exitosamente (CON CORRECCIÓN DE COMBOS):")
        print(f"   Productos: {resumen['total_productos']}")
        print(f"   Unidades REALES: {resumen['total_unidades_reales']:g}")
        print(f"   Total: ${resumen['total_ventas']:.2f}")
        print(f"   Incluye corrección de combos: SÍ")
        
        return jsonify({
            'success': True,
            'productos': productos,
            'resumen': resumen,
            'parametros': {
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'categoria': categoria,
                'orden': orden,
                'solo_con_ventas': solo_con_ventas,
                'incluye_todas_facturas': True,
                'correccion_combos_aplicada': True
            }
        })
        
    except Exception as e:
        print(f"❌ Error en reporte de ventas: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/exportar_reporte_ventas')
def exportar_reporte_ventas():
    """Exportar reporte de ventas a Excel o CSV - CORREGIDO PARA COMBOS"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros (mismos que el reporte)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        categoria = request.args.get('categoria', '').strip()
        orden = request.args.get('orden', 'cantidad_desc')
        formato = request.args.get('formato', 'csv')  # csv o excel
        
        # Validar fechas
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        
        print(f"📤 Exportando reporte a {formato.upper()}: {fecha_desde} a {fecha_hasta}")
        
        # *** QUERY CORREGIDA (MISMA DEL REPORTE) ***
        query = db.session.query(
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo,
            # *** CANTIDAD REAL CALCULADA ***
            func.sum(
                case(
                    (Producto.es_combo == True, DetalleFactura.cantidad * Producto.cantidad_combo),
                    else_=DetalleFactura.cantidad
                )
            ).label('cantidad_real_vendida'),
            func.sum(DetalleFactura.cantidad).label('unidades_combos_vendidas'),
            func.sum(DetalleFactura.subtotal).label('total_vendido'),
            func.avg(DetalleFactura.precio_unitario).label('precio_promedio'),
            func.max(Factura.fecha).label('ultima_venta'),
            func.count(DetalleFactura.id).label('num_transacciones')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt
                # *** SIN FILTRO DE ESTADO - Incluye todas las facturas ***
            )
        )
        
        if categoria:
            query = query.filter(Producto.categoria == categoria)
        
        query = query.group_by(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo
        )
        
        # *** ORDENAMIENTO USANDO CANTIDAD REAL ***
        if orden == 'cantidad_desc':
            query = query.order_by(desc('cantidad_real_vendida'))
        elif orden == 'cantidad_asc':
            query = query.order_by(asc('cantidad_real_vendida'))
        elif orden == 'total_desc':
            query = query.order_by(desc('total_vendido'))
        elif orden == 'total_asc':
            query = query.order_by(asc('total_vendido'))
        elif orden == 'codigo':
            query = query.order_by(Producto.codigo)
        elif orden == 'nombre':
            query = query.order_by(Producto.nombre)
        
        resultados = query.all()
        print(f"📊 Exportando {len(resultados)} productos")
        
        # *** PREPARAR DATOS PARA EXPORTACIÓN CON INFORMACIÓN DE COMBOS ***
        datos_exportacion = []
        encabezados = [
            'Código',
            'Producto',
            'Descripción',
            'Categoría',
            'Tipo',  # *** NUEVO: Producto Base / Combo
            'Cantidad Real Vendida',  # *** CORREGIDO ***
            'Unidades/Combos',  # *** NUEVO: Detalle de combos ***
            'Precio Promedio',
            'Total Vendido',
            'Última Venta',
            'Número de Transacciones'
        ]
        
        # Agregar encabezados
        datos_exportacion.append(encabezados)
        
        # *** AGREGAR DATOS CON INFORMACIÓN DE COMBOS ***
        for resultado in resultados:
            # Calcular cantidad real y información de combo
            cantidad_real = float(resultado.cantidad_real_vendida) if resultado.cantidad_real_vendida else 0.0
            unidades_combos = int(resultado.unidades_combos_vendidas) if resultado.unidades_combos_vendidas else 0
            
            # Información del tipo de producto
            if resultado.es_combo:
                tipo_producto = "Combo/Oferta"
                cantidad_combo = float(resultado.cantidad_combo) if resultado.cantidad_combo else 1.0
                detalle_unidades = f"{unidades_combos} combos × {cantidad_combo:g} c/u"
            else:
                tipo_producto = "Producto Base"
                detalle_unidades = f"{int(cantidad_real)} unidades"
            
            fila = [
                resultado.codigo,
                resultado.nombre,
                resultado.descripcion or '',
                resultado.categoria or 'Sin categoría',
                tipo_producto,  # *** NUEVO ***
                f"{cantidad_real:g}",  # *** CANTIDAD REAL ***
                detalle_unidades,  # *** DETALLE ***
                f"{float(resultado.precio_promedio):.2f}" if resultado.precio_promedio else "0.00",
                f"{float(resultado.total_vendido):.2f}" if resultado.total_vendido else "0.00",
                resultado.ultima_venta.strftime('%d/%m/%Y') if resultado.ultima_venta else 'N/A',
                int(resultado.num_transacciones) if resultado.num_transacciones else 0
            ]
            datos_exportacion.append(fila)
        
        # Generar archivo según formato
        if formato == 'excel':
            return generar_excel_reporte(datos_exportacion, fecha_desde, fecha_hasta)
        else:  # CSV por defecto
            return generar_csv_reporte(datos_exportacion, fecha_desde, fecha_hasta)
        
    except Exception as e:
        print(f"❌ Error exportando reporte: {str(e)}")
        return jsonify({'error': f'Error al exportar: {str(e)}'}), 500
        

def generar_csv_reporte(datos, fecha_desde, fecha_hasta):
    """Generar archivo CSV del reporte"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Escribir encabezado del reporte
    writer.writerow([f'Reporte de Ventas por Producto'])
    writer.writerow([f'Período: {fecha_desde} al {fecha_hasta}'])
    writer.writerow([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    writer.writerow([])  # Línea vacía
    
    # Escribir datos
    for fila in datos:
        writer.writerow(fila)
    
    # Crear respuesta
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.csv'
    
    return response

def generar_excel_reporte(datos, fecha_desde, fecha_hasta):
    """Generar archivo Excel del reporte (requiere openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        # Estilos
        titulo_font = Font(bold=True, size=16)
        encabezado_font = Font(bold=True, size=12, color="FFFFFF")
        encabezado_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Título del reporte
        ws['A1'] = f'Reporte de Ventas por Producto'
        ws['A1'].font = titulo_font
        ws['A2'] = f'Período: {fecha_desde} al {fecha_hasta}'
        ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Fila inicial para datos
        fila_inicio = 5
        
        # Escribir encabezados
        encabezados = datos[0]
        for col, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal='center')
        
        # Escribir datos
        for fila_idx, fila_datos in enumerate(datos[1:], fila_inicio + 1):
            for col_idx, valor in enumerate(fila_datos, 1):
                ws.cell(row=fila_idx, column=col_idx, value=valor)
        
        # Ajustar ancho de columnas
        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.xlsx'
        
        return response
        
    except ImportError:
        # Si no está instalado openpyxl, devolver CSV
        return generar_csv_reporte(datos, fecha_desde, fecha_hasta)

# ==================== REPORTE RÁPIDO DE TOP PRODUCTOS ====================

@app.route('/api/top_productos_vendidos')
def api_top_productos_vendidos():
    """API para obtener top 10 productos más vendidos (últimos 30 días)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Últimos 30 días
        fecha_hasta = datetime.now()
        fecha_desde = fecha_hasta - timedelta(days=30)
        
        # Query para top productos
        resultados = db.session.query(
            Producto.codigo,
            Producto.nombre,
            func.sum(DetalleFactura.cantidad).label('cantidad_vendida'),
            func.sum(DetalleFactura.subtotal).label('total_vendido')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde,
                Factura.fecha <= fecha_hasta,
                Factura.estado == 'autorizada'
            )
        ).group_by(
            Producto.id,
            Producto.codigo,
            Producto.nombre
        ).order_by(
            desc('cantidad_vendida')
        ).limit(10).all()
        
        # Formatear respuesta
        top_productos = []
        for resultado in resultados:
            top_productos.append({
                'codigo': resultado.codigo,
                'nombre': resultado.nombre,
                'cantidad_vendida': int(resultado.cantidad_vendida),
                'total_vendido': float(resultado.total_vendido)
            })
        
        return jsonify({
            'success': True,
            'productos': top_productos,
            'periodo': '30 días'
        })
        
    except Exception as e:
        print(f"Error en top productos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== DASHBOARD DE VENTAS PARA INCLUIR EN MAIN ====================

# REEMPLAZA tu función api_dashboard_ventas() existente con esta versión mejorada:

@app.route('/api/dashboard_ventas')
def api_dashboard_ventas():
    """API para dashboard de ventas (resumen del día) - VERSIÓN CORREGIDA"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        print("🔍 Iniciando consulta dashboard ventas...")
        
        # Obtener fecha actual
        from datetime import date
        hoy = date.today()
        print(f"📅 Consultando ventas para: {hoy}")
        
        # CONSULTA 1: Datos básicos de ventas del día
        # Usar DATE() para comparar solo la fecha, no hora
        consulta_ventas = db.session.query(
            func.count(Factura.id).label('num_facturas'),
            func.coalesce(func.sum(Factura.total), 0).label('total_vendido')
        ).filter(
            func.date(Factura.fecha) == hoy
        ).first()
        
        print(f"📊 Consulta ventas básicas completada")
        print(f"   Facturas: {consulta_ventas.num_facturas}")
        print(f"   Total: ${consulta_ventas.total_vendido}")
        
        # CONSULTA 2: Total de unidades vendidas del día
        consulta_unidades = db.session.query(
            func.coalesce(func.sum(DetalleFactura.cantidad), 0).label('total_unidades')
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            func.date(Factura.fecha) == hoy
        ).first()
        
        print(f"📦 Unidades vendidas: {consulta_unidades.total_unidades}")
        
        # CONSULTA 3: Producto más vendido del día
        consulta_top_producto = db.session.query(
            Producto.codigo,
            Producto.nombre,
            func.sum(DetalleFactura.cantidad).label('cantidad_vendida')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            func.date(Factura.fecha) == hoy
        ).group_by(
            Producto.id,
            Producto.codigo, 
            Producto.nombre
        ).order_by(
            desc('cantidad_vendida')
        ).first()
        
        if consulta_top_producto:
            print(f"👑 Top producto: {consulta_top_producto.codigo} - {consulta_top_producto.nombre} ({consulta_top_producto.cantidad_vendida} unidades)")
        else:
            print("👑 No hay ventas de productos hoy")
        
        # Preparar respuesta
        response_data = {
            'success': True,
            'ventas_hoy': {
                'num_facturas': int(consulta_ventas.num_facturas or 0),
                'total_vendido': float(consulta_ventas.total_vendido or 0),
                'unidades_vendidas': int(consulta_unidades.total_unidades or 0)
            },
            'producto_top_hoy': None
        }
        
        # Agregar producto top si existe
        if consulta_top_producto:
            response_data['producto_top_hoy'] = {
                'codigo': consulta_top_producto.codigo,
                'nombre': consulta_top_producto.nombre,
                'cantidad': int(consulta_top_producto.cantidad_vendida)
            }
        
        print(f"✅ Dashboard data preparada correctamente")
        print(f"📤 Enviando respuesta: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Error en api_dashboard_ventas: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        
        # Devolver datos por defecto en caso de error
        return jsonify({
            'success': True,
            'ventas_hoy': {
                'num_facturas': 0,
                'total_vendido': 0.0,
                'unidades_vendidas': 0
            },
            'producto_top_hoy': None,
            'error_debug': str(e)
        })

# AGREGAR TAMBIÉN ESTA RUTA PARA DEBUG:
@app.route('/debug/dashboard_data')
def debug_dashboard_data():
    """Endpoint para debugging del dashboard"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        from datetime import date
        hoy = date.today()
        
        # Información de debug
        debug_info = {
            'fecha_hoy': str(hoy),
            'total_facturas_bd': Factura.query.count(),
            'total_productos_bd': Producto.query.count(),
            'facturas_hoy': Factura.query.filter(func.date(Factura.fecha) == hoy).count(),
            'ultimas_facturas': []
        }
        
        # Últimas 5 facturas para debug
        ultimas_facturas = Factura.query.order_by(Factura.id.desc()).limit(5).all()
        for factura in ultimas_facturas:
            debug_info['ultimas_facturas'].append({
                'id': factura.id,
                'numero': factura.numero,
                'total': float(factura.total),
                'fecha': factura.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                'estado': factura.estado
            })
        
        # Probar la consulta de dashboard
        try:
            dashboard_data = api_dashboard_ventas()
            debug_info['dashboard_response'] = dashboard_data.get_json()
        except Exception as e:
            debug_info['dashboard_error'] = str(e)
        
        return jsonify({
            'success': True,
            'debug_info': debug_info
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== PASO 3: FUNCIÓN PARA MIGRAR DATOS EXISTENTES ====================

def migrar_detalle_facturas_con_iva():
    """Migrar detalles de facturas existentes para agregar IVA individual"""
    try:
        print("🔄 Iniciando migración de detalles con IVA...")
        
        # Buscar detalles sin porcentaje_iva o importe_iva
        detalles_sin_iva = DetalleFactura.query.filter(
            or_(
                DetalleFactura.porcentaje_iva.is_(None),
                DetalleFactura.importe_iva.is_(None),
                DetalleFactura.porcentaje_iva == 0
            )
        ).all()
        
        print(f"📋 Encontrados {len(detalles_sin_iva)} detalles para migrar")
        
        contador_migrados = 0
        
        for detalle in detalles_sin_iva:
            try:
                # Obtener porcentaje de IVA del producto
                if detalle.producto:
                    porcentaje_iva = float(detalle.producto.iva)
                else:
                    porcentaje_iva = 21.0  # Por defecto
                
                # Calcular importe de IVA
                subtotal = float(detalle.subtotal)
                importe_iva = round((subtotal * porcentaje_iva / 100), 2)
                
                # Actualizar campos
                detalle.porcentaje_iva = Decimal(str(porcentaje_iva))
                detalle.importe_iva = Decimal(str(importe_iva))
                
                contador_migrados += 1
                
                if contador_migrados % 50 == 0:
                    print(f"   📊 Migrados {contador_migrados}/{len(detalles_sin_iva)}")
                
            except Exception as e:
                print(f"⚠️ Error migrando detalle ID {detalle.id}: {e}")
        
        # Guardar cambios
        if contador_migrados > 0:
            db.session.commit()
            print(f"✅ Migración completada: {contador_migrados} detalles actualizados")
        else:
            print("✅ No hay detalles para migrar")
            
        return contador_migrados
        
    except Exception as e:
        print(f"❌ Error en migración: {e}")
        db.session.rollback()
        return 0


# ==================== PASO 4: FUNCIÓN PARA VERIFICAR IVA POR DETALLE ====================

@app.route('/api/verificar_iva_detalle/<int:factura_id>')
def verificar_iva_detalle(factura_id):
    """Verificar que los detalles tengan IVA individual correcto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        detalles_info = []
        total_iva_calculado = 0
        
        for detalle in factura.detalles:
            # IVA guardado en BD
            iva_bd = float(detalle.importe_iva) if detalle.importe_iva else 0
            porcentaje_bd = float(detalle.porcentaje_iva) if detalle.porcentaje_iva else 0
            
            # IVA recalculado
            subtotal = float(detalle.subtotal)
            porcentaje_producto = float(detalle.producto.iva) if detalle.producto else 21.0
            iva_recalculado = round((subtotal * porcentaje_producto / 100), 2)
            
            total_iva_calculado += iva_recalculado
            
            detalle_info = {
                'id': detalle.id,
                'producto': detalle.producto.nombre if detalle.producto else 'Sin producto',
                'subtotal': subtotal,
                'porcentaje_bd': porcentaje_bd,
                'porcentaje_producto': porcentaje_producto,
                'iva_bd': iva_bd,
                'iva_recalculado': iva_recalculado,
                'coincide': abs(iva_bd - iva_recalculado) < 0.01
            }
            
            detalles_info.append(detalle_info)
        
        # Comparar con total de factura
        iva_factura = float(factura.iva)
        
        return jsonify({
            'success': True,
            'factura_numero': factura.numero,
            'iva_factura': iva_factura,
            'iva_calculado_suma': round(total_iva_calculado, 2),
            'total_coincide': abs(iva_factura - total_iva_calculado) < 0.01,
            'detalles': detalles_info
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== PASO 5: RUTA PARA MIGRACIÓN MANUAL ====================

@app.route('/migrar_iva_detalles', methods=['POST'])
def migrar_iva_detalles_endpoint():
    """Endpoint para ejecutar migración de IVA en detalles"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        contador = migrar_detalle_facturas_con_iva()
        
        return jsonify({
            'success': True,
            'mensaje': f'Migración completada: {contador} detalles actualizados',
            'detalles_migrados': contador
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en migración: {str(e)}'
        }), 500

@app.route('/importar_productos')
def importar_productos_vista():
    """Vista para importar productos desde Excel"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('importar_productos.html')

@app.route('/api/importar_productos_lote', methods=['POST'])
def importar_productos_lote():
    """Importar un lote de productos desde Excel"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        productos = data.get('productos', [])
        opciones = data.get('opciones', {})
        
        solo_actualizar = opciones.get('solo_actualizar', False)
        crear_nuevos = opciones.get('crear_nuevos', True)
        
        resultados = {
            'nuevos': 0,
            'actualizados': 0,
            'errores': 0,
            'detalles_errores': [],
            'productos_procesados': []  # NUEVO: Array con el estado de cada producto
        }
        
        print(f"📦 Procesando lote de {len(productos)} productos...")
        
        for producto_data in productos:
            producto_resultado = {
                'codigo': producto_data.get('codigo', 'UNKNOWN'),
                'estado': 'error',  # Por defecto error, se cambia si todo va bien
                'mensaje': ''
            }
            
            try:
                codigo = str(producto_data['codigo']).strip().upper()
                descripcion = str(producto_data['descripcion']).strip()
                precio = float(producto_data['precio'])
                
                producto_resultado['codigo'] = codigo  # Actualizar con código limpio
                
                # Validaciones básicas
                if not codigo or not descripcion or precio <= 0:
                    producto_resultado['estado'] = 'error'
                    producto_resultado['mensaje'] = 'Datos inválidos'
                    resultados['errores'] += 1
                    resultados['detalles_errores'].append({
                        'error': f'Datos inválidos para código {codigo}',
                        'codigo': codigo
                    })
                    resultados['productos_procesados'].append(producto_resultado)
                    continue
                
                # Buscar si el producto ya existe
                producto_existente = Producto.query.filter_by(codigo=codigo).first()
                
                if producto_existente:
                    # Producto existe - actualizar si está permitido
                    if solo_actualizar:
                        # Actualizar precio y descripción
                        producto_existente.precio = Decimal(str(precio))
                        producto_existente.nombre = descripcion
                        producto_existente.descripcion = descripcion
                        producto_existente.fecha_modificacion = datetime.now()
                        
                        # Si no tiene costo, intentar calcularlo con margen por defecto
                        if not producto_existente.costo or producto_existente.costo == 0:
                            margen_defecto = 30.0  # 30% de margen por defecto
                            costo_calculado = precio / (1 + (margen_defecto / 100))
                            producto_existente.costo = Decimal(str(costo_calculado))
                            producto_existente.margen = Decimal(str(margen_defecto))
                        
                        # ACTUALIZAR ESTADO
                        producto_resultado['estado'] = 'actualizado'
                        producto_resultado['mensaje'] = 'Producto actualizado correctamente'
                        
                        resultados['actualizados'] += 1
                        print(f"✅ Actualizado: {codigo} - {descripcion}")
                    else:
                        # Si no se permite actualizar, contar como error
                        producto_resultado['estado'] = 'error'
                        producto_resultado['mensaje'] = 'Producto ya existe (actualización deshabilitada)'
                        
                        resultados['errores'] += 1
                        resultados['detalles_errores'].append({
                            'error': f'Producto {codigo} ya existe (actualización deshabilitada)',
                            'codigo': codigo
                        })
                        
                else:
                    # Producto nuevo - crear si está permitido
                    if crear_nuevos:
                        # Calcular costo con margen por defecto del 30%
                        margen_defecto = 30.0
                        costo_calculado = precio / (1 + (margen_defecto / 100))
                        
                        # Detectar categoría básica desde la descripción
                        categoria = detectar_categoria(descripcion)
                        
                        nuevo_producto = Producto(
                            codigo=codigo,
                            nombre=descripcion,
                            descripcion=descripcion,
                            precio=Decimal(str(precio)),                    # <- Del Excel
                            costo=Decimal(str(float(producto_data.get('costo', 0)))),   # <- Del Excel
                            margen=Decimal(str(float(producto_data.get('margen', 30)))), # <- Del Excel
                            stock=0,  # Stock inicial en 0
                            categoria=categoria,
                            iva=Decimal('21.0'),  # IVA por defecto 21%
                            activo=True,
                            es_combo=False,
                            fecha_creacion=datetime.now(),
                            fecha_modificacion=datetime.now()
                        )
                        
                        db.session.add(nuevo_producto)
                        
                        # ACTUALIZAR ESTADO
                        producto_resultado['estado'] = 'nuevo'
                        producto_resultado['mensaje'] = 'Producto creado correctamente'
                        
                        resultados['nuevos'] += 1
                        print(f"🆕 Creado: {codigo} - {descripcion}")
                    else:
                        # Si no se permite crear nuevos, contar como error
                        producto_resultado['estado'] = 'error'
                        producto_resultado['mensaje'] = 'Producto no existe (creación deshabilitada)'
                        
                        resultados['errores'] += 1
                        resultados['detalles_errores'].append({
                            'error': f'Producto {codigo} no existe (creación deshabilitada)',
                            'codigo': codigo
                        })
                
            except Exception as e:
                producto_resultado['estado'] = 'error'
                producto_resultado['mensaje'] = f'Error: {str(e)}'
                
                resultados['errores'] += 1
                resultados['detalles_errores'].append({
                    'error': f'Error procesando {producto_data.get("codigo", "UNKNOWN")}: {str(e)}',
                    'codigo': producto_data.get('codigo', 'UNKNOWN')
                })
                print(f"❌ Error en producto {producto_data.get('codigo', 'UNKNOWN')}: {str(e)}")
            
            # SIEMPRE agregar el resultado del producto al array
            resultados['productos_procesados'].append(producto_resultado)
        
        # Confirmar cambios en la base de datos
        db.session.commit()
        
        print(f"✅ Lote completado: {resultados['nuevos']} nuevos, {resultados['actualizados']} actualizados, {resultados['errores']} errores")
        
        return jsonify({
            'success': True,
            'message': 'Lote procesado correctamente',
            'nuevos': resultados['nuevos'],
            'actualizados': resultados['actualizados'],
            'errores': resultados['errores'],
            'detalles_errores': resultados['detalles_errores'],
            'productos_procesados': resultados['productos_procesados']  # NUEVO: Estados detallados
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en importación de lote: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al procesar lote: {str(e)}'
        }), 500


def detectar_categoria(descripcion):
    """Detectar categoría básica desde la descripción del producto"""
    descripcion_lower = descripcion.lower()
    
    # Mapeo de palabras clave a categorías
    categorias = {
        'POLLO': ['pollo', 'pechuga', 'muslo', 'ala', 'carcasa'],
        'CARNE': ['carne', 'bife', 'asado', 'costilla', 'vacio', 'chorizo'],
        'CERDO': ['cerdo', 'bondiola', 'matambre', 'costilla cerdo'],
        'PESCADO': ['pescado', 'salmon', 'merluza', 'atun'],
        'CHACINADOS': ['salame', 'jamon', 'mortadela', 'chorizo', 'morcilla'],
        'LACTEOS': ['leche', 'queso', 'yogur', 'manteca', 'crema'],
        'CONGELADOS': ['congelado', 'frozen', 'helado'],
        'BEBIDAS': ['gaseosa', 'agua', 'jugo', 'cerveza', 'vino'],
        'PANADERIA': ['pan', 'facturas', 'torta', 'galletas'],
        'LIMPIEZA': ['detergente', 'lavandina', 'jabon', 'shampoo'],
        'VERDURAS': ['verdura', 'lechuga', 'tomate', 'cebolla', 'papa']
    }
    
    for categoria, palabras_clave in categorias.items():
        if any(palabra in descripcion_lower for palabra in palabras_clave):
            return categoria
    
    return 'GENERAL'  # Categoría por defecto

# AGREGAR esta nueva ruta en app.py:

@app.route('/api/eliminar_combo/<int:combo_id>', methods=['DELETE'])
def eliminar_combo(combo_id):
    """Eliminar un combo con validaciones de seguridad"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Verificar que el producto existe y es un combo
        combo = Producto.query.get_or_404(combo_id)
        
        if not combo.es_combo:
            return jsonify({
                'success': False,
                'error': 'El producto no es un combo'
            }), 400
        
        print(f"🗑️ Solicitud de eliminación para combo: {combo.codigo} - {combo.nombre}")
        
        # VALIDACIÓN 1: Verificar si el combo se vendió alguna vez
        ventas_combo = DetalleFactura.query.filter_by(producto_id=combo_id).count()
        
        if ventas_combo > 0:
            print(f"❌ Combo {combo.codigo} tiene {ventas_combo} ventas registradas")
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar el combo porque tiene {ventas_combo} ventas registradas. Solo puedes desactivarlo.',
                'motivo': 'tiene_ventas',
                'ventas_count': ventas_combo,
                'sugerencia': 'Usa el botón de desactivar en lugar de eliminar'
            }), 400
        
        # VALIDACIÓN 2: Verificar si está en facturas pendientes o con errores
        facturas_pendientes = db.session.query(DetalleFactura).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                DetalleFactura.producto_id == combo_id,
                or_(
                    Factura.estado == 'pendiente',
                    Factura.estado == 'error_afip'
                )
            )
        ).count()
        
        if facturas_pendientes > 0:
            print(f"❌ Combo {combo.codigo} está en {facturas_pendientes} facturas pendientes")
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar el combo porque está en {facturas_pendientes} facturas pendientes de autorización.',
                'motivo': 'facturas_pendientes',
                'facturas_count': facturas_pendientes
            }), 400
        
        # VALIDACIÓN 3: Verificar si tiene stock (opcional - puedes eliminarlo igual)
        if combo.stock > 0:
            print(f"⚠️ Advertencia: Combo {combo.codigo} tiene stock {combo.stock}")
            # No bloquear, solo advertir
        
        # SI LLEGAMOS AQUÍ: Es seguro eliminar
        print(f"✅ Combo {combo.codigo} puede eliminarse de forma segura")
        
        # Eliminar el combo de la base de datos
        nombre_combo = combo.nombre
        codigo_combo = combo.codigo
        
        db.session.delete(combo)
        db.session.commit()
        
        print(f"🗑️ Combo eliminado exitosamente: {codigo_combo}")
        
        return jsonify({
            'success': True,
            'message': f'Combo "{nombre_combo}" eliminado correctamente',
            'codigo_eliminado': codigo_combo,
            'motivo': 'eliminacion_segura'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error eliminando combo: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno al eliminar combo: {str(e)}'
        }), 500


@app.route('/api/verificar_eliminacion_combo/<int:combo_id>')
def verificar_eliminacion_combo(combo_id):
    """Verificar si un combo puede ser eliminado de forma segura"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        combo = Producto.query.get_or_404(combo_id)
        
        if not combo.es_combo:
            return jsonify({
                'puede_eliminar': False,
                'motivo': 'no_es_combo'
            })
        
        # Verificar ventas
        ventas_count = DetalleFactura.query.filter_by(producto_id=combo_id).count()
        
        # Verificar facturas pendientes
        facturas_pendientes = db.session.query(DetalleFactura).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                DetalleFactura.producto_id == combo_id,
                or_(
                    Factura.estado == 'pendiente',
                    Factura.estado == 'error_afip'
                )
            )
        ).count()
        
        puede_eliminar = (ventas_count == 0 and facturas_pendientes == 0)
        
        resultado = {
            'puede_eliminar': puede_eliminar,
            'ventas_count': ventas_count,
            'facturas_pendientes': facturas_pendientes,
            'stock': combo.stock,
            'codigo': combo.codigo,
            'nombre': combo.nombre
        }
        
        if not puede_eliminar:
            if ventas_count > 0:
                resultado['motivo'] = 'tiene_ventas'
                resultado['mensaje'] = f'El combo tiene {ventas_count} ventas registradas'
            elif facturas_pendientes > 0:
                resultado['motivo'] = 'facturas_pendientes'
                resultado['mensaje'] = f'El combo está en {facturas_pendientes} facturas pendientes'
        else:
            resultado['motivo'] = 'puede_eliminar'
            resultado['mensaje'] = 'El combo puede eliminarse de forma segura'
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'puede_eliminar': False,
            'error': str(e)
        }), 500

# AGREGAR estas rutas en app.py:

@app.route('/api/buscar_facturas')
def buscar_facturas():
    """Buscar facturas con filtros avanzados"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros de búsqueda
        numero = request.args.get('numero', '').strip()
        cliente = request.args.get('cliente', '').strip()
        estado = request.args.get('estado', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        limite = int(request.args.get('limite', 100))  # Limitar resultados
        
        print(f"🔍 Búsqueda de facturas:")
        print(f"   Número: '{numero}'")
        print(f"   Cliente: '{cliente}'")
        print(f"   Estado: '{estado}'")
        print(f"   Fecha desde: '{fecha_desde}'")
        print(f"   Fecha hasta: '{fecha_hasta}'")
        print(f"   Límite: {limite}")
        
        # Construir query base con join a cliente
        query = db.session.query(Factura).join(Cliente, Factura.cliente_id == Cliente.id)
        
        # Aplicar filtros
        if numero:
            query = query.filter(Factura.numero.ilike(f'%{numero}%'))
            print(f"   Filtro aplicado: Número contiene '{numero}'")
        
        if cliente:
            query = query.filter(Cliente.nombre.ilike(f'%{cliente}%'))
            print(f"   Filtro aplicado: Cliente contiene '{cliente}'")
        
        if estado:
            query = query.filter(Factura.estado == estado)
            print(f"   Filtro aplicado: Estado = '{estado}'")
        
        # Filtros de fecha
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
                query = query.filter(Factura.fecha >= fecha_desde_dt)
                print(f"   Filtro aplicado: Fecha >= {fecha_desde}")
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha desde inválido. Use YYYY-MM-DD'
                }), 400
        
        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                query = query.filter(Factura.fecha <= fecha_hasta_dt)
                print(f"   Filtro aplicado: Fecha <= {fecha_hasta}")
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha hasta inválido. Use YYYY-MM-DD'
                }), 400
        
        # Ordenar por fecha descendente (más recientes primero)
        query = query.order_by(Factura.fecha.desc())
        
        # Aplicar límite
        facturas = query.limit(limite).all()
        
        print(f"   Facturas encontradas: {len(facturas)}")
        
        # Formatear resultados
        resultado = []
        for factura in facturas:
            # ✅ BUSCAR DESCUENTO APLICADO
            descuento = DescuentoFactura.query.filter_by(factura_id=factura.id).first()

            # Obtener información de medios de pago
            medios_pago = []
            for medio in factura.medios_pago:
                medios_pago.append({
                    'medio_pago': medio.medio_pago,
                    'importe': float(medio.importe)
                })
            
            factura_dict = {
                'id': factura.id,
                'numero': factura.numero,
                'fecha': factura.fecha.strftime('%d/%m/%Y %H:%M'),
                'fecha_iso': factura.fecha.isoformat(),
                'cliente': {
                    'id': factura.cliente.id,
                    'nombre': factura.cliente.nombre,
                    'documento': factura.cliente.documento,
                    'tipo_documento': factura.cliente.tipo_documento
                },
                'tipo_comprobante': factura.tipo_comprobante,
                'tipo_comprobante_nombre': obtener_nombre_comprobante(factura.tipo_comprobante),
                'subtotal': float(factura.subtotal),
                'iva': float(factura.iva),
                'total': float(factura.total),
                'estado': factura.estado,
                'estado_descripcion': obtener_descripcion_estado(factura.estado),
                'cae': factura.cae,
                'vto_cae': factura.vto_cae.strftime('%d/%m/%Y') if factura.vto_cae else None,
                'medios_pago': medios_pago,
                'cantidad_items': len(factura.detalles),
                'usuario': factura.usuario.nombre if factura.usuario else 'Desconocido',
                 'tiene_descuento': bool(descuento),
                'descuento_porcentaje': float(descuento.porcentaje_descuento) if descuento else 0,
                'descuento_monto': float(descuento.monto_descuento) if descuento else 0,
                'total_original': float(descuento.total_original) if descuento else float(factura.total)
            }
            
            resultado.append(factura_dict)
        
        return jsonify({
            'success': True,
            'facturas': resultado,
            'total': len(resultado),
            'limite_aplicado': len(facturas) == limite,
            'filtros_aplicados': {
                'numero': numero,
                'cliente': cliente,
                'estado': estado,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta
            }
        })
        
    except Exception as e:
        print(f"❌ Error buscando facturas: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'Error en la búsqueda: {str(e)}'
        }), 500


def obtener_nombre_comprobante(tipo):
    """Obtener nombre legible del tipo de comprobante"""
    tipos = {
        '01': 'Factura A',
        '06': 'Factura B', 
        '11': 'Factura C',
        '51': 'Factura M'
    }
    return tipos.get(str(tipo), f'Tipo {tipo}')


def obtener_descripcion_estado(estado):
    """Obtener descripción legible del estado"""
    estados = {
        'autorizada': 'Autorizada por AFIP',
        'pendiente': 'Pendiente de autorización',
        'error_afip': 'Error en AFIP',
        'anulada': 'Anulada'
    }
    return estados.get(estado, estado.title())


@app.route('/api/reintentar_afip/<int:factura_id>', methods=['POST'])
def reintentar_afip(factura_id):
    """Reintentar autorización AFIP para una factura pendiente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        if factura.estado not in ['pendiente', 'error_afip']:
            return jsonify({
                'success': False,
                'error': f'No se puede reintentar. Estado actual: {factura.estado}'
            }), 400
        
        print(f"🔄 Reintentando autorización AFIP para factura {factura.numero}")
        
        # Preparar datos para AFIP
        cliente = factura.cliente
        
        # Obtener items con IVA detallado
        items_detalle = []
        for detalle in factura.detalles:
            items_detalle.append({
                'subtotal': float(detalle.subtotal),
                'iva_porcentaje': float(detalle.porcentaje_iva) if detalle.porcentaje_iva else 21.0
            })
        
        datos_comprobante = {
            'tipo_comprobante': int(factura.tipo_comprobante),
            'punto_venta': factura.punto_venta,
            'importe_neto': float(factura.subtotal),
            'importe_iva': float(factura.iva),
            'items_detalle': items_detalle,
            'doc_tipo': 99,  # Sin identificar por defecto
            'doc_nro': 0
        }
        
        # Agregar datos del cliente si existen
        if cliente and cliente.documento:
            if cliente.tipo_documento == 'CUIT' and len(cliente.documento) == 11:
                datos_comprobante['doc_tipo'] = 80  # CUIT
                datos_comprobante['doc_nro'] = int(cliente.documento)
            elif cliente.tipo_documento == 'DNI' and len(cliente.documento) >= 7:
                datos_comprobante['doc_tipo'] = 96  # DNI
                datos_comprobante['doc_nro'] = int(cliente.documento)
        
        # Intentar autorizar en AFIP
        resultado_afip = arca_client.autorizar_comprobante(datos_comprobante)
        
        if resultado_afip['success']:
            # Actualizar factura con datos de AFIP
            numero_afip = resultado_afip['numero']
            
            # Verificar si el número de AFIP ya existe
            factura_existente = Factura.query.filter(
                and_(Factura.numero == numero_afip, Factura.id != factura.id)
            ).first()
            
            if not factura_existente:
                factura.numero = numero_afip
            
            factura.cae = resultado_afip['cae']
            factura.vto_cae = resultado_afip['vto_cae']
            factura.estado = 'autorizada'
            
            db.session.commit()
            
            print(f"✅ Reintento exitoso. CAE: {factura.cae}")
            
            return jsonify({
                'success': True,
                'message': f'Factura autorizada exitosamente',
                'numero': factura.numero,
                'cae': factura.cae,
                'estado': factura.estado
            })
        else:
            # Actualizar estado a error
            factura.estado = 'error_afip'
            db.session.commit()
            
            print(f"❌ Reintento falló: {resultado_afip.get('error', 'Error desconocido')}")
            
            return jsonify({
                'success': False,
                'error': f"Error AFIP: {resultado_afip.get('error', 'Error desconocido')}",
                'estado': factura.estado
            })
        
    except Exception as e:
        print(f"❌ Error en reintento AFIP: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500


@app.route('/api/anular_factura/<int:factura_id>', methods=['POST'])
def anular_factura(factura_id):
    """Anular una factura (marcar como anulada)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        if factura.estado == 'anulada':
            return jsonify({
                'success': False,
                'error': 'La factura ya está anulada'
            }), 400
        
        motivo = request.json.get('motivo', '').strip() if request.json else ''
        
        print(f"❌ Anulando factura {factura.numero}. Motivo: {motivo}")
        
        # Marcar como anulada
        factura.estado = 'anulada'
        
        # TODO: Aquí podrías agregar una tabla de motivos de anulación
        # y restaurar stock si es necesario
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Factura {factura.numero} anulada correctamente',
            'estado': factura.estado
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error anulando factura: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al anular factura: {str(e)}'
        }), 500


# PASO 2: Nuevas rutas en app.py para gestionar acceso rápido

@app.route('/api/toggle_acceso_rapido/<int:producto_id>', methods=['POST'])
def toggle_acceso_rapido(producto_id):
    """Activar/desactivar producto como acceso rápido"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        if producto.acceso_rapido:
            # Quitar de acceso rápido
            producto.acceso_rapido = False
            producto.orden_acceso_rapido = 0
            accion = 'removido de'
        else:
            # Verificar límite de 8 productos
            productos_acceso_rapido = Producto.query.filter_by(
                acceso_rapido=True, 
                activo=True
            ).count()
            
            if productos_acceso_rapido >= 8:
                return jsonify({
                    'success': False,
                    'error': 'Máximo 8 productos permitidos en acceso rápido. Quita uno primero.'
                }), 400
            
            # Agregar a acceso rápido
            producto.acceso_rapido = True
            # Asignar orden automáticamente
            max_orden = db.session.query(func.max(Producto.orden_acceso_rapido)).scalar() or 0
            producto.orden_acceso_rapido = max_orden + 1
            accion = 'agregado a'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Producto {accion} acceso rápido',
            'acceso_rapido': producto.acceso_rapido,
            'orden': producto.orden_acceso_rapido
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error en toggle_acceso_rapido: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al cambiar acceso rápido: {str(e)}'
        }), 500


@app.route('/api/reordenar_acceso_rapido', methods=['POST'])
def reordenar_acceso_rapido():
    """Reordenar productos de acceso rápido"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        productos_orden = data.get('productos_orden', [])
        
        # productos_orden debe ser una lista de IDs en el orden deseado
        for i, producto_id in enumerate(productos_orden):
            producto = Producto.query.get(producto_id)
            if producto and producto.acceso_rapido:
                producto.orden_acceso_rapido = i + 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Orden actualizado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error al reordenar: {str(e)}'
        }), 500


@app.route('/api/productos_acceso_rapido')
def obtener_productos_acceso_rapido():
    """Obtener productos marcados como acceso rápido"""
    try:
        productos = Producto.query.filter_by(
            acceso_rapido=True,
            activo=True
        ).order_by(
            Producto.orden_acceso_rapido.asc(),
            Producto.codigo.asc()
        ).limit(8).all()
        
        productos_data = []
        for producto in productos:
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio': float(producto.precio),
                'stock': producto.stock,
                'iva': float(producto.iva),
                'orden': producto.orden_acceso_rapido
            })
        
        return jsonify({
            'success': True,
            'productos': productos_data,
            'total': len(productos_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



def migrar_acceso_rapido():
    """Migrar productos existentes para agregar funcionalidad de acceso rápido"""
    try:
        print("🔄 Migrando productos para acceso rápido...")
        
        # Tomar los primeros 8 productos activos como acceso rápido inicial
        productos_iniciales = Producto.query.filter_by(activo=True).limit(8).all()
        
        for i, producto in enumerate(productos_iniciales):
            producto.acceso_rapido = True
            producto.orden_acceso_rapido = i + 1
            print(f"  ✅ {producto.codigo} marcado como acceso rápido (orden {i + 1})")
        
        # Marcar el resto como NO acceso rápido
        productos_restantes = Producto.query.filter(
            ~Producto.id.in_([p.id for p in productos_iniciales])
        ).all()
        
        for producto in productos_restantes:
            producto.acceso_rapido = False
            producto.orden_acceso_rapido = 0
        
        db.session.commit()
        print(f"✅ Migración completada: {len(productos_iniciales)} productos en acceso rápido")
        
    except Exception as e:
        print(f"❌ Error en migración: {e}")
        db.session.rollback()


# PASO 5: Ruta para ejecutar migración
@app.route('/migrar_acceso_rapido', methods=['POST'])
def ejecutar_migracion_acceso_rapido():
    """Endpoint para ejecutar migración de acceso rápido"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        migrar_acceso_rapido()
        
        return jsonify({
            'success': True,
            'mensaje': 'Migración de acceso rápido completada'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en migración: {str(e)}'
        }), 500


@app.route('/reportes')
def reportes():
    return render_template('reportes.html')

from flask import render_template

@app.route("/reporte_ventas_parcial")
def reporte_ventas_parcial():
    return render_template("reporte_ventas_parcial.html")

# ==================== RUTAS DE GASTOS ====================
# Agregar estas rutas al final de tu app.py antes del if __name__ == '__main__':

@app.route('/api/gastos', methods=['GET'])
def obtener_gastos():
    """Obtener gastos filtrados por fechas"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
        
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')
        
        # Validar parámetros
        if not fecha_desde or not fecha_hasta:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar fecha_desde y fecha_hasta'
            }), 400
        
        # Validar formato de fechas
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        print(f"📋 Obteniendo gastos del {fecha_desde} al {fecha_hasta}")
        
        # Consulta principal
        gastos = Gasto.query.filter(
            and_(
                Gasto.fecha >= fecha_desde_dt,
                Gasto.fecha <= fecha_hasta_dt,
                Gasto.activo == True
            )
        ).order_by(Gasto.fecha.desc(), Gasto.fecha_creacion.desc()).all()
        
        # Convertir a lista de diccionarios
        gastos_list = [gasto.to_dict() for gasto in gastos]
        
        # Calcular totales por categoría usando el método del modelo
        estadisticas = Gasto.calcular_gastos_por_fecha(fecha_desde_dt, fecha_hasta_dt)
        categorias_dict = estadisticas['gastos_por_categoria'] if estadisticas else {}
        
        # Calcular total general
        total_gastos = sum(gasto['monto'] for gasto in gastos_list)
        
        print(f"✅ Gastos obtenidos: {len(gastos_list)} gastos (${total_gastos:.2f})")
        
        return jsonify({
            'success': True,
            'gastos': gastos_list,
            'categorias': categorias_dict,
            'resumen': {
                'total_gastos': total_gastos,
                'cantidad_gastos': len(gastos_list),
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'promedio_por_gasto': total_gastos / len(gastos_list) if gastos_list else 0
            }
        })
        
    except Exception as e:
        print(f"❌ Error al obtener gastos: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos', methods=['POST'])
def crear_gasto():
    """Crear un nuevo gasto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = ['descripcion', 'monto', 'fecha']
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                }), 400
        
        # Validar monto
        try:
            monto = float(data['monto'])
            if monto <= 0:
                return jsonify({
                    'success': False,
                    'error': 'El monto debe ser mayor a 0'
                }), 400
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Monto inválido'
            }), 400
        
        # Validar fecha
        try:
            fecha_obj = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        # Datos del gasto
        descripcion = data['descripcion'].strip()
        categoria = data.get('categoria', 'general')
        metodo_pago = data.get('metodo_pago', 'efectivo')
        notas = data.get('notas', '').strip()
        
        # Validar longitud de campos
        if len(descripcion) > 200:
            return jsonify({
                'success': False,
                'error': 'La descripción no puede exceder 200 caracteres'
            }), 400
        
        # Crear nuevo gasto
        gasto = Gasto(
            fecha=fecha_obj,
            descripcion=descripcion,
            monto=Decimal(str(monto)),
            categoria=categoria,
            metodo_pago=metodo_pago,
            notas=notas if notas else None,
            usuario_id=session['user_id']
        )
        
        db.session.add(gasto)
        db.session.commit()
        
        print(f"✅ Gasto creado: ID {gasto.id} - {descripcion} - ${monto:.2f}")
        
        return jsonify({
            'success': True,
            'message': 'Gasto registrado exitosamente',
            'gasto': gasto.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al crear gasto: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/<int:gasto_id>', methods=['DELETE'])
def eliminar_gasto(gasto_id):
    """Eliminar un gasto (soft delete)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        gasto = Gasto.query.get_or_404(gasto_id)
        
        if not gasto.activo:
            return jsonify({
                'success': False,
                'error': 'El gasto ya está eliminado'
            }), 404
        
        # Hacer soft delete
        gasto.activo = False
        gasto.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        print(f"✅ Gasto eliminado: ID {gasto_id} - {gasto.descripcion}")
        
        return jsonify({
            'success': True,
            'message': 'Gasto eliminado exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al eliminar gasto: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/<int:gasto_id>', methods=['PUT'])
def actualizar_gasto(gasto_id):
    """Actualizar un gasto existente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.get_json()
        gasto = Gasto.query.get_or_404(gasto_id)
        
        if not gasto.activo:
            return jsonify({
                'success': False,
                'error': 'No se puede actualizar un gasto eliminado'
            }), 400
        
        # Actualizar campos proporcionados
        if 'descripcion' in data:
            descripcion = data['descripcion'].strip()
            if not descripcion:
                return jsonify({
                    'success': False,
                    'error': 'La descripción no puede estar vacía'
                }), 400
            if len(descripcion) > 200:
                return jsonify({
                    'success': False,
                    'error': 'La descripción no puede exceder 200 caracteres'
                }), 400
            gasto.descripcion = descripcion
        
        if 'monto' in data:
            try:
                monto = float(data['monto'])
                if monto <= 0:
                    return jsonify({
                        'success': False,
                        'error': 'El monto debe ser mayor a 0'
                    }), 400
                gasto.monto = Decimal(str(monto))
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Monto inválido'
                }), 400
        
        if 'fecha' in data:
            try:
                gasto.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha inválido'
                }), 400
        
        if 'categoria' in data:
            gasto.categoria = data['categoria']
        
        if 'metodo_pago' in data:
            gasto.metodo_pago = data['metodo_pago']
        
        if 'notas' in data:
            notas = data['notas'].strip()
            gasto.notas = notas if notas else None
        
        gasto.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        print(f"✅ Gasto actualizado: ID {gasto_id}")
        
        return jsonify({
            'success': True,
            'message': 'Gasto actualizado exitosamente',
            'gasto': gasto.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al actualizar gasto: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/categorias', methods=['GET'])
def obtener_categorias_gastos():
    """Obtener categorías de gastos disponibles y estadísticas de uso"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener categorías disponibles
        categorias_disponibles = Gasto.obtener_categorias_disponibles()
        
        # Obtener estadísticas de uso de categorías
        categorias_usadas = db.session.query(
            Gasto.categoria,
            func.count(Gasto.id).label('cantidad'),
            func.sum(Gasto.monto).label('total')
        ).filter(
            Gasto.activo == True
        ).group_by(Gasto.categoria).order_by(desc('total')).all()
        
        # Combinar información
        categorias_con_stats = []
        for cat_info in categorias_disponibles:
            codigo = cat_info['codigo']
            
            # Buscar estadísticas para esta categoría
            stats = next((cat for cat in categorias_usadas if cat[0] == codigo), None)
            
            categoria_completa = {
                'codigo': codigo,
                'nombre': cat_info['nombre'],
                'cantidad_gastos': stats[1] if stats else 0,
                'total_gastado': float(stats[2]) if stats and stats[2] else 0.0,
                'en_uso': bool(stats)
            }
            
            categorias_con_stats.append(categoria_completa)
        
        return jsonify({
            'success': True,
            'categorias': categorias_con_stats
        })
        
    except Exception as e:
        print(f"❌ Error al obtener categorías: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/medios_pago', methods=['GET'])
def obtener_medios_pago_gastos():
    """Obtener medios de pago disponibles para gastos"""
    try:
        # Usar los mismos medios de pago que las facturas
        medios_disponibles = MedioPago.obtener_medios_disponibles()
        
        return jsonify({
            'success': True,
            'medios_pago': medios_disponibles
        })
        
    except Exception as e:
        print(f"❌ Error al obtener medios de pago: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/resumen_periodo')
def resumen_gastos_periodo():
    """Obtener resumen de gastos para un período específico"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')
        
        if not fecha_desde or not fecha_hasta:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar fechas desde y hasta'
            }), 400
        
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        
        # Obtener resumen usando método del modelo
        resumen = Gasto.calcular_gastos_por_fecha(fecha_desde_dt, fecha_hasta_dt)
        
        if not resumen:
            return jsonify({
                'success': False,
                'error': 'Error al calcular resumen'
            }), 500
        
        # Obtener gastos por medio de pago
        gastos_por_medio = Gasto.obtener_gastos_por_medio_pago(fecha_desde_dt, fecha_hasta_dt)
        
        # Obtener gasto promedio por día
        dias_periodo = (fecha_hasta_dt - fecha_desde_dt).days + 1
        promedio_diario = resumen['total_general'] / dias_periodo if dias_periodo > 0 else 0
        
        return jsonify({
            'success': True,
            'resumen': {
                'total_general': resumen['total_general'],
                'gastos_por_categoria': resumen['gastos_por_categoria'],
                'gastos_por_medio_pago': gastos_por_medio,
                'periodo': {
                    'desde': fecha_desde,
                    'hasta': fecha_hasta,
                    'dias': dias_periodo
                },
                'promedios': {
                    'diario': round(promedio_diario, 2)
                }
            }
        })
        
    except Exception as e:
        print(f"❌ Error en resumen de período: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/test_conexion', methods=['GET'])
def test_conexion_gastos():
    """Probar que la tabla de gastos funciona correctamente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Contar gastos activos
        total_gastos = Gasto.query.filter_by(activo=True).count()
        
        # Obtener último gasto
        ultimo_gasto = Gasto.query.filter_by(activo=True).order_by(Gasto.id.desc()).first()
        
        return jsonify({
            'success': True,
            'message': 'Conexión con gastos exitosa',
            'total_gastos': total_gastos,
            'ultimo_gasto': ultimo_gasto.to_dict() if ultimo_gasto else None,
            'categorias_disponibles': len(Gasto.obtener_categorias_disponibles())
        })
        
    except Exception as e:
        print(f"❌ Error test conexión gastos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/producto_precio_oferta/<int:producto_id>/<float:cantidad>')
def obtener_precio_con_oferta_api(producto_id, cantidad):
    """API para obtener precio con oferta aplicada"""
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        precio_final = producto.obtener_precio_con_oferta(cantidad)
        info_oferta = producto.obtener_info_oferta(cantidad)
        
        return jsonify({
            'success': True,
            'precio_base': float(producto.precio),
            'precio_final': precio_final,
            'cantidad': cantidad,
            'info_oferta': info_oferta,
            'producto': {
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'stock': producto.stock
            }
        })
        
    except Exception as e:
        print(f"Error en API precio oferta: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ofertas_producto/<int:producto_id>')
def obtener_ofertas_producto(producto_id):
    """Obtener todas las ofertas de un producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        ofertas = OfertaVolumen.query.filter_by(
            producto_id=producto_id,
            activo=True
        ).order_by(OfertaVolumen.cantidad_minima.asc()).all()
        
        ofertas_data = [oferta.to_dict() for oferta in ofertas]
        
        return jsonify({
            'success': True,
            'producto': {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio_base': float(producto.precio)
            },
            'ofertas': ofertas_data,
            'tiene_ofertas': len(ofertas_data) > 0
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/crear_oferta_volumen', methods=['POST'])
def crear_oferta_volumen():
    """Crear nueva oferta por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        producto_id = data.get('producto_id')
        cantidad_minima = float(data.get('cantidad_minima', 0))
        precio_oferta = float(data.get('precio_oferta', 0))
        descripcion = data.get('descripcion', '').strip()
        
        # Validaciones
        if not producto_id:
            return jsonify({'error': 'Producto requerido'}), 400
        
        if cantidad_minima <= 0:
            return jsonify({'error': 'La cantidad mínima debe ser mayor a 0'}), 400
        
        if precio_oferta <= 0:
            return jsonify({'error': 'El precio de oferta debe ser mayor a 0'}), 400
        
        # Verificar que el producto existe
        producto = Producto.query.get_or_404(producto_id)
        
        # Verificar que el precio de oferta sea menor al precio normal
        if precio_oferta >= float(producto.precio):
            return jsonify({
                'error': f'El precio de oferta (${precio_oferta}) debe ser menor al precio normal (${producto.precio})'
            }), 400
        
        # Verificar que no exista una oferta igual
        oferta_existente = OfertaVolumen.query.filter_by(
            producto_id=producto_id,
            cantidad_minima=cantidad_minima,
            activo=True
        ).first()
        
        if oferta_existente:
            return jsonify({
                'error': f'Ya existe una oferta para cantidad mínima {cantidad_minima}'
            }), 400
        
        # Crear oferta
        nueva_oferta = OfertaVolumen(
            producto_id=producto_id,
            cantidad_minima=Decimal(str(cantidad_minima)),
            precio_oferta=Decimal(str(precio_oferta)),
            descripcion=descripcion if descripcion else None
        )
        
        db.session.add(nueva_oferta)
        db.session.commit()
        
        print(f"Oferta creada: {producto.codigo} - {cantidad_minima}+ = ${precio_oferta}")
        
        return jsonify({
            'success': True,
            'message': f'Oferta creada: desde {cantidad_minima} unidades a ${precio_oferta}',
            'oferta': nueva_oferta.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error creando oferta: {e}")
        return jsonify({'error': f'Error al crear oferta: {str(e)}'}), 500

############# RUTAS DE OFERTAS
@app.route('/api/eliminar_oferta_volumen/<int:oferta_id>', methods=['DELETE'])
def eliminar_oferta_volumen(oferta_id):
    """Eliminar oferta por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        oferta = OfertaVolumen.query.get_or_404(oferta_id)
        
        # Marcar como inactiva en lugar de eliminar
        oferta.activo = False
        oferta.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Oferta eliminada correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al eliminar oferta: {str(e)}'}), 500

# ==================== RUTAS ADICIONALES PARA OFERTAS POR VOLUMEN ====================
# Agregar estas rutas después de las rutas existentes de ofertas en tu app.py

@app.route('/ofertas_volumen')
def ofertas_volumen():
    """Vista principal para gestionar ofertas por volumen"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('ofertas_volumen.html')


@app.route('/api/ofertas_volumen_todas')
def obtener_todas_ofertas_volumen():
    """Obtener todas las ofertas por volumen con información del producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Consulta con JOIN para obtener ofertas con información del producto
        ofertas = db.session.query(OfertaVolumen).join(
            Producto, OfertaVolumen.producto_id == Producto.id
        ).filter(
            Producto.activo == True  # Solo productos activos
        ).order_by(
            OfertaVolumen.fecha_creacion.desc()
        ).all()
        
        ofertas_data = []
        for oferta in ofertas:
            oferta_dict = oferta.to_dict()
            
            # Agregar información completa del producto
            if oferta.producto:
                oferta_dict['producto'] = {
                    'id': oferta.producto.id,
                    'codigo': oferta.producto.codigo,
                    'nombre': oferta.producto.nombre,
                    'precio': float(oferta.producto.precio),
                    'stock': oferta.producto.stock,
                    'categoria': oferta.producto.categoria,
                    'activo': oferta.producto.activo
                }
            else:
                oferta_dict['producto'] = None
            
            ofertas_data.append(oferta_dict)
        
        print(f"📊 Devolviendo {len(ofertas_data)} ofertas por volumen")
        
        return jsonify({
            'success': True,
            'ofertas': ofertas_data,
            'total': len(ofertas_data)
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo ofertas: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/actualizar_oferta_volumen/<int:oferta_id>', methods=['PUT'])
def actualizar_oferta_volumen(oferta_id):
    """Actualizar una oferta por volumen existente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # Validaciones
        cantidad_minima = float(data.get('cantidad_minima', 0))
        precio_oferta = float(data.get('precio_oferta', 0))
        descripcion = data.get('descripcion', '').strip()
        
        if cantidad_minima <= 0:
            return jsonify({'error': 'La cantidad mínima debe ser mayor a 0'}), 400
        
        if precio_oferta <= 0:
            return jsonify({'error': 'El precio de oferta debe ser mayor a 0'}), 400
        
        # Obtener oferta existente
        oferta = OfertaVolumen.query.get_or_404(oferta_id)
        
        # Verificar que el precio de oferta sea menor al precio normal
        if precio_oferta >= float(oferta.producto.precio):
            return jsonify({
                'error': f'El precio de oferta (${precio_oferta}) debe ser menor al precio normal (${oferta.producto.precio})'
            }), 400
        
        # Verificar que no exista otra oferta igual (excluyendo la actual)
        oferta_duplicada = OfertaVolumen.query.filter(
            and_(
                OfertaVolumen.producto_id == oferta.producto_id,
                OfertaVolumen.cantidad_minima == cantidad_minima,
                OfertaVolumen.id != oferta_id,
                OfertaVolumen.activo == True
            )
        ).first()
        
        if oferta_duplicada:
            return jsonify({
                'error': f'Ya existe otra oferta para cantidad mínima {cantidad_minima}'
            }), 400
        
        # Actualizar oferta
        oferta.cantidad_minima = Decimal(str(cantidad_minima))
        oferta.precio_oferta = Decimal(str(precio_oferta))
        oferta.descripcion = descripcion if descripcion else None
        oferta.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        print(f"Oferta actualizada: {oferta.producto.codigo} - {cantidad_minima}+ = ${precio_oferta}")
        
        return jsonify({
            'success': True,
            'message': f'Oferta actualizada: desde {cantidad_minima} unidades a ${precio_oferta}',
            'oferta': oferta.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error actualizando oferta: {e}")
        return jsonify({'error': f'Error al actualizar oferta: {str(e)}'}), 500


@app.route('/api/ofertas_activas_resumen')
def obtener_resumen_ofertas_activas():
    """Obtener resumen de ofertas activas para dashboard"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Contar ofertas activas
        ofertas_activas = OfertaVolumen.query.filter_by(activo=True).count()
        
        # Contar productos con ofertas
        productos_con_ofertas = db.session.query(OfertaVolumen.producto_id).filter_by(activo=True).distinct().count()
        
        # Obtener descuento promedio
        ofertas_con_descuento = db.session.query(
            OfertaVolumen.precio_oferta,
            Producto.precio
        ).join(
            Producto, OfertaVolumen.producto_id == Producto.id
        ).filter(
            OfertaVolumen.activo == True
        ).all()
        
        descuento_promedio = 0
        if ofertas_con_descuento:
            descuentos = []
            for precio_oferta, precio_normal in ofertas_con_descuento:
                if float(precio_normal) > 0:
                    descuento = ((float(precio_normal) - float(precio_oferta)) / float(precio_normal)) * 100
                    descuentos.append(descuento)
            
            if descuentos:
                descuento_promedio = sum(descuentos) / len(descuentos)
        
        return jsonify({
            'success': True,
            'ofertas_activas': ofertas_activas,
            'productos_con_ofertas': productos_con_ofertas,
            'descuento_promedio': round(descuento_promedio, 1)
        })
        
    except Exception as e:
        print(f"Error obteniendo resumen de ofertas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/validar_oferta_volumen', methods=['POST'])
def validar_oferta_volumen():
    """Validar datos de una oferta antes de crearla/actualizarla"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        producto_id = data.get('producto_id')
        cantidad_minima = float(data.get('cantidad_minima', 0))
        precio_oferta = float(data.get('precio_oferta', 0))
        oferta_id = data.get('oferta_id')  # Para edición
        
        errores = []
        
        # Validar producto
        if not producto_id:
            errores.append('Debe seleccionar un producto')
        else:
            producto = Producto.query.get(producto_id)
            if not producto:
                errores.append('Producto no encontrado')
            elif not producto.activo:
                errores.append('El producto está inactivo')
            else:
                # Validar precio
                if precio_oferta >= float(producto.precio):
                    errores.append(f'El precio de oferta debe ser menor a ${float(producto.precio):.2f}')
                
                # Verificar duplicados
                query = OfertaVolumen.query.filter(
                    and_(
                        OfertaVolumen.producto_id == producto_id,
                        OfertaVolumen.cantidad_minima == cantidad_minima,
                        OfertaVolumen.activo == True
                    )
                )
                
                if oferta_id:  # Excluir la oferta actual si es edición
                    query = query.filter(OfertaVolumen.id != oferta_id)
                
                if query.first():
                    errores.append(f'Ya existe una oferta para cantidad mínima {cantidad_minima}')
        
        # Validaciones básicas
        if cantidad_minima <= 0:
            errores.append('La cantidad mínima debe ser mayor a 0')
        
        if precio_oferta <= 0:
            errores.append('El precio de oferta debe ser mayor a 0')
        
        # Calcular información de la oferta si es válida
        info_oferta = None
        if not errores and producto:
            precio_normal = float(producto.precio)
            descuento_porcentaje = ((precio_normal - precio_oferta) / precio_normal) * 100
            ahorro_total = (precio_normal - precio_oferta) * cantidad_minima
            
            info_oferta = {
                'precio_normal': precio_normal,
                'precio_oferta': precio_oferta,
                'cantidad_minima': cantidad_minima,
                'descuento_porcentaje': round(descuento_porcentaje, 1),
                'ahorro_por_unidad': round(precio_normal - precio_oferta, 2),
                'ahorro_total': round(ahorro_total, 2)
            }
        
        return jsonify({
            'success': len(errores) == 0,
            'errores': errores,
            'info_oferta': info_oferta
        })
        
    except Exception as e:
        print(f"Error validando oferta: {e}")
        return jsonify({
            'success': False,
            'errores': [f'Error de validación: {str(e)}']
        }), 500


@app.route('/api/productos_sin_ofertas')
def obtener_productos_sin_ofertas():
    """Obtener productos que no tienen ofertas por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Productos que NO tienen ofertas activas
        productos_con_ofertas = db.session.query(OfertaVolumen.producto_id).filter_by(activo=True).distinct().subquery()
        
        productos_sin_ofertas = Producto.query.filter(
            and_(
                Producto.activo == True,
                ~Producto.id.in_(productos_con_ofertas)
            )
        ).order_by(Producto.codigo).limit(20).all()
        
        productos_data = []
        for producto in productos_sin_ofertas:
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio': float(producto.precio),
                'stock': producto.stock,
                'categoria': producto.categoria
            })
        
        return jsonify({
            'success': True,
            'productos': productos_data,
            'mensaje': f'Se encontraron {len(productos_data)} productos sin ofertas'
        })
        
    except Exception as e:
        print(f"Error obteniendo productos sin ofertas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/estadisticas_ofertas_volumen')
def obtener_estadisticas_ofertas():
    """Obtener estadísticas detalladas de las ofertas por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Estadísticas básicas
        total_ofertas = OfertaVolumen.query.count()
        ofertas_activas = OfertaVolumen.query.filter_by(activo=True).count()
        ofertas_inactivas = total_ofertas - ofertas_activas
        
        # Productos con ofertas
        productos_con_ofertas = db.session.query(
            OfertaVolumen.producto_id
        ).filter_by(activo=True).distinct().count()
        
        # Ofertas por rango de descuento
        ofertas_con_productos = db.session.query(
            OfertaVolumen.precio_oferta,
            Producto.precio.label('precio_normal')
        ).join(
            Producto, OfertaVolumen.producto_id == Producto.id
        ).filter(
            OfertaVolumen.activo == True
        ).all()
        
        rangos_descuento = {
            'menos_10': 0,      # Menos del 10%
            'entre_10_20': 0,   # Entre 10% y 20%
            'entre_20_30': 0,   # Entre 20% y 30%
            'mas_30': 0         # Más del 30%
        }
        
        descuentos = []
        for precio_oferta, precio_normal in ofertas_con_productos:
            if float(precio_normal) > 0:
                descuento = ((float(precio_normal) - float(precio_oferta)) / float(precio_normal)) * 100
                descuentos.append(descuento)
                
                if descuento < 10:
                    rangos_descuento['menos_10'] += 1
                elif descuento < 20:
                    rangos_descuento['entre_10_20'] += 1
                elif descuento < 30:
                    rangos_descuento['entre_20_30'] += 1
                else:
                    rangos_descuento['mas_30'] += 1
        
        # Calcular estadísticas de descuentos
        descuento_promedio = sum(descuentos) / len(descuentos) if descuentos else 0
        descuento_minimo = min(descuentos) if descuentos else 0
        descuento_maximo = max(descuentos) if descuentos else 0
        
        # Top 5 productos con más ofertas
        top_productos = db.session.query(
            Producto.codigo,
            Producto.nombre,
            func.count(OfertaVolumen.id).label('cantidad_ofertas')
        ).join(
            OfertaVolumen, Producto.id == OfertaVolumen.producto_id
        ).filter(
            OfertaVolumen.activo == True
        ).group_by(
            Producto.id, Producto.codigo, Producto.nombre
        ).order_by(
            func.count(OfertaVolumen.id).desc()
        ).limit(5).all()
        
        top_productos_data = []
        for codigo, nombre, cantidad in top_productos:
            top_productos_data.append({
                'codigo': codigo,
                'nombre': nombre,
                'cantidad_ofertas': cantidad
            })
        
        return jsonify({
            'success': True,
            'estadisticas': {
                'totales': {
                    'total_ofertas': total_ofertas,
                    'ofertas_activas': ofertas_activas,
                    'ofertas_inactivas': ofertas_inactivas,
                    'productos_con_ofertas': productos_con_ofertas
                },
                'descuentos': {
                    'promedio': round(descuento_promedio, 1),
                    'minimo': round(descuento_minimo, 1),
                    'maximo': round(descuento_maximo, 1),
                    'rangos': rangos_descuento
                },
                'top_productos': top_productos_data
            }
        })
        
    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/productos_con_ofertas_volumen')
def obtener_productos_con_ofertas_volumen():
    """Obtener todos los productos que tienen ofertas por volumen activas"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener productos con ofertas por volumen
        productos_con_ofertas = db.session.query(Producto).join(
            OfertaVolumen, Producto.id == OfertaVolumen.producto_id
        ).filter(
            and_(
                Producto.activo == True,
                OfertaVolumen.activo == True
            )
        ).distinct().all()
        
        resultado = {}
        
        for producto in productos_con_ofertas:
            ofertas = OfertaVolumen.query.filter_by(
                producto_id=producto.id,
                activo=True
            ).order_by(OfertaVolumen.cantidad_minima.asc()).all()
            
            resultado[str(producto.id)] = {
                'producto': {
                    'id': producto.id,
                    'codigo': producto.codigo,
                    'nombre': producto.nombre,
                    'precio': float(producto.precio),
                    'precio_base': float(producto.precio)
                },
                'ofertas': [oferta.to_dict() for oferta in ofertas]
            }
        
        return jsonify({
            'success': True,
            'productos_ofertas': resultado,
            'total_productos': len(resultado)
        })
        
    except Exception as e:
        print(f"Error obteniendo productos con ofertas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Agregar esta ruta en tu archivo principal de Flask (app.py)

@app.route('/estadisticas')
def estadisticas():
    """Página de estadísticas y reportes avanzados"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        return render_template('estadisticas.html')
    except Exception as e:
        flash(f'Error cargando estadísticas: {str(e)}')
        return redirect(url_for('index'))


@app.route('/api/descuento_factura/<int:factura_id>')
def obtener_descuento_factura(factura_id):
    """Obtener información del descuento aplicado a una factura"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        descuento = DescuentoFactura.query.filter_by(factura_id=factura_id).first()
        
        #print(f"🔍 DEBUG API: Buscando descuento para factura {factura_id}")
        #print(f"🔍 DEBUG API: Descuento encontrado: {bool(descuento)}")
        
        if descuento:
            #print(f"🔍 DEBUG API: Descuento: {descuento.porcentaje_descuento}% = ${descuento.monto_descuento}")
            return jsonify({
                'success': True,
                'tiene_descuento': True,
                'descuento': descuento.to_dict()
            })
        else:
            return jsonify({
                'success': True,
                'tiene_descuento': False
            })
    except Exception as e:
       # print(f"❌ DEBUG API: Error: {e}")
        return jsonify({'success': False, 'error': str(e)})

app.run(debug=True, host='0.0.0.0', port=5080)